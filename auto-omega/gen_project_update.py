"""
gen_project_update.py — Tạo bảng kế hoạch thay đổi Project-Update.xlsx
từ phân tích 2 file feedback khách hàng:
  - Web Omega - Feedback.docx
  - Web Omega - BS nội dung.docx
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent / "Project-Update.xlsx"

# ── Màu sắc ────────────────────────────────────────────────────────────────────
C_HEADER_BG   = "0D5C38"   # xanh đậm omega
C_HEADER_FG   = "FFFFFF"
C_P1_BG       = "FF4444"   # đỏ — Rất cao
C_P2_BG       = "FF9900"   # cam — Cao
C_P3_BG       = "FFD700"   # vàng — Trung bình
C_P4_BG       = "A8D5A2"   # xanh lá nhạt — Thấp
C_ROW_ODD     = "F5FBF7"
C_ROW_EVEN    = "FFFFFF"
C_SECTION_BG  = "E8F5E9"   # xanh lá rất nhạt cho dòng phân nhóm

PRIORITY_COLORS = {
    "🔴 Rất cao": "FF4444",
    "🟠 Cao":     "FF9900",
    "🟡 Trung bình": "FFD700",
    "🟢 Thấp":    "A8D5A2",
}

# ── Dữ liệu ───────────────────────────────────────────────────────────────────
# Cột: STT | Hạng mục | Mô tả thay đổi | File(s) bị tác động | Độ ưu tiên | Loại | Nguồn | Ghi chú/Đánh giá
CHANGES = [
    # ══════════════════════ NHÓM 1: THAY ĐỔI VĂN BẢN / THƯƠNG HIỆU ══════════════════════
    {
        "group":    "NHÓM 1 — Thay đổi tên gọi & thương hiệu",
        "stt":      1,
        "hang_muc": "Đổi tên thương hiệu trong Giải pháp",
        "mo_ta":    '"OMEGA ERP" → "OMEGA" trên toàn bộ trang giai-phap.html và 8 trang giải pháp ngành',
        "files":    "giai-phap.html\ngiai-phap/nong-nghiep.html\ngiai-phap/xay-dung.html\ngiai-phap/thuong-mai.html\ngiai-phap/san-xuat.html\ngiai-phap/bao-hiem.html\ngiai-phap/y-te.html\ngiai-phap/giao-duc.html\ngiai-phap/dich-vu.html",
        "priority": "🔴 Rất cao",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Thay đổi định vị thương hiệu — ảnh hưởng SEO & nhận diện. Cần tìm-thay toàn bộ cụm 'OMEGA ERP' → 'OMEGA' (cẩn thận với context).",
    },
    {
        "group":    None,
        "stt":      2,
        "hang_muc": "Đổi tên trong trang Dịch vụ",
        "mo_ta":    '"Omega ERP" → "Phần mềm của Omega" trên trang dich-vu.html',
        "files":    "dich-vu.html",
        "priority": "🔴 Rất cao",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Đồng bộ với thay đổi thương hiệu ở mục 1.",
    },
    {
        "group":    None,
        "stt":      3,
        "hang_muc": "Xóa 'ERP' khỏi banner Giải pháp",
        "mo_ta":    'Xóa chữ "ERP" khỏi banner/hero section trang giai-phap.html',
        "files":    "giai-phap.html",
        "priority": "🔴 Rất cao",
        "loai":     "Text + UI",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Banner là phần nhìn thấy đầu tiên — ưu tiên cao.",
    },
    {
        "group":    None,
        "stt":      4,
        "hang_muc": "Đổi tên section Hệ sinh thái — Trang chủ",
        "mo_ta":    'Đổi tên section trang chủ → "Hệ sinh thái giải pháp công nghệ Omega"',
        "files":    "index.html",
        "priority": "🔴 Rất cao",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Thay đổi ngắn gọn, tác động lớn đến định vị thương hiệu.",
    },
    # ══════════════════════ NHÓM 2: CẬP NHẬT NỘI DUNG TRANG CHỦ ══════════════════════
    {
        "group":    "NHÓM 2 — Trang chủ (index.html)",
        "stt":      5,
        "hang_muc": "Logo — đổi màu gradient xanh",
        "mo_ta":    "Cập nhật logo trang chủ sang màu xanh gradient theo đặc tả thiết kế mới",
        "files":    "index.html\nassets/images/logo*",
        "priority": "🟠 Cao",
        "loai":     "UI + Assets",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Cần file logo mới từ designer. Nếu chỉ CSS thì chỉnh filter/SVG gradient.",
    },
    {
        "group":    None,
        "stt":      6,
        "hang_muc": "Animation background tối hơn",
        "mo_ta":    "Tăng độ tối của background animation hero section trang chủ",
        "files":    "index.html\nassets/css/omega.css",
        "priority": "🟠 Cao",
        "loai":     "UI/CSS",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Chỉnh giá trị opacity/brightness trong CSS. Nhanh.",
    },
    {
        "group":    None,
        "stt":      7,
        "hang_muc": "Tagline & hero subtitle trang chủ",
        "mo_ta":    "Cập nhật tagline chính và subtitle trong hero section theo nội dung mới từ BS nội dung.docx",
        "files":    "index.html",
        "priority": "🟠 Cao",
        "loai":     "Text",
        "nguon":    "Feedback.docx + BS nội dung.docx",
        "ghi_chu":  "Cần xác nhận đúng text từ file BS nội dung. Ảnh hưởng SEO (H1, meta description).",
    },
    {
        "group":    None,
        "stt":      8,
        "hang_muc": "Section headers trang chủ",
        "mo_ta":    "Cập nhật tiêu đề các section theo yêu cầu feedback",
        "files":    "index.html",
        "priority": "🟠 Cao",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Xem lại từng section header, so sánh với BS nội dung để thống nhất.",
    },
    # ══════════════════════ NHÓM 3: VỀ OMEGA ══════════════════════
    {
        "group":    "NHÓM 3 — Trang Về Omega (ve-omega.html)",
        "stt":      9,
        "hang_muc": "Câu chuyện người sáng lập",
        "mo_ta":    'Thêm mục "Câu chuyện người sáng lập" — ông Nguyễn Văn Nhân, lý do đặt tên OMEGA',
        "files":    "ve-omega.html",
        "priority": "🟠 Cao",
        "loai":     "Nội dung mới",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Nội dung branding quan trọng — tạo sự khác biệt & niềm tin. Cần section mới trong HTML.",
    },
    {
        "group":    None,
        "stt":      10,
        "hang_muc": "Ý nghĩa logo Omega",
        "mo_ta":    "Thêm giải thích ý nghĩa biểu tượng logo Omega vào trang Về Omega",
        "files":    "ve-omega.html",
        "priority": "🟠 Cao",
        "loai":     "Nội dung mới",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Có thể kết hợp với câu chuyện sáng lập (mục 9) thành 1 section.",
    },
    {
        "group":    None,
        "stt":      11,
        "hang_muc": '"Điểm mạnh khác biệt" → "Giá trị khác biệt"',
        "mo_ta":    'Đổi tên section từ "Điểm mạnh khác biệt" sang "Giá trị khác biệt"',
        "files":    "ve-omega.html",
        "priority": "🟡 Trung bình",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Thay đổi nhỏ nhưng cần đồng bộ nếu có anchor link.",
    },
    {
        "group":    None,
        "stt":      12,
        "hang_muc": "Tầm nhìn Top 5 VN 2030",
        "mo_ta":    "Cập nhật nội dung tầm nhìn: Top 5 doanh nghiệp phần mềm ERP Việt Nam đến 2030",
        "files":    "ve-omega.html",
        "priority": "🟡 Trung bình",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Ảnh hưởng định vị chiến lược — cần xác nhận số liệu chính xác.",
    },
    {
        "group":    None,
        "stt":      13,
        "hang_muc": "Sứ mệnh — cập nhật văn bản",
        "mo_ta":    "Cập nhật đoạn văn sứ mệnh theo nội dung mới từ feedback",
        "files":    "ve-omega.html",
        "priority": "🟡 Trung bình",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Đồng bộ với slide thuyết trình / tài liệu nội bộ.",
    },
    {
        "group":    None,
        "stt":      14,
        "hang_muc": "5 Giá trị cốt lõi — full text",
        "mo_ta":    "Cập nhật nội dung đầy đủ 5 giá trị: Tận tâm, Đồng hành, Gắn kết, Trách nhiệm, Đổi mới",
        "files":    "ve-omega.html",
        "priority": "🟡 Trung bình",
        "loai":     "Text",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Hiện tại chỉ có tên, cần thêm đoạn mô tả chi tiết cho mỗi giá trị.",
    },
    {
        "group":    None,
        "stt":      15,
        "hang_muc": "Animation Về Omega — thay đổi",
        "mo_ta":    "Cập nhật animation section Về Omega theo yêu cầu thiết kế mới",
        "files":    "ve-omega.html\nassets/css/omega.css",
        "priority": "🟢 Thấp",
        "loai":     "UI/CSS",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Cần spec rõ hơn về loại animation mong muốn trước khi implement.",
    },
    {
        "group":    None,
        "stt":      16,
        "hang_muc": "Đổi tên section sản phẩm trong Về Omega",
        "mo_ta":    "Đổi tên section giới thiệu sản phẩm trong trang Về Omega theo tên mới",
        "files":    "ve-omega.html",
        "priority": "🟢 Thấp",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Xác nhận tên section mới từ BS nội dung.",
    },
    # ══════════════════════ NHÓM 4: SẢN PHẨM ══════════════════════
    {
        "group":    "NHÓM 4 — Trang Sản phẩm (san-pham.html)",
        "stt":      17,
        "hang_muc": "Cấu trúc nhóm sản phẩm mới",
        "mo_ta":    "Tái cấu trúc danh mục: Tài chính-Kế toán | Nhân sự-Tiền lương | Sản xuất | Mua hàng-Bán hàng | Nền tảng",
        "files":    "san-pham.html",
        "priority": "🔴 Rất cao",
        "loai":     "Cấu trúc + UI",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Thay đổi lớn — ảnh hưởng layout tổng thể trang sản phẩm. Cần thiết kế lại section.",
    },
    {
        "group":    None,
        "stt":      18,
        "hang_muc": "Thêm module mới: OMEGA.EM",
        "mo_ta":    "Thêm module Quản lý Bảo trì (OMEGA.EM) với 7 chức năng con vào danh sách sản phẩm",
        "files":    "san-pham.html",
        "priority": "🔴 Rất cao",
        "loai":     "Nội dung mới",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Module hoàn toàn mới — cần card mới + trang chi tiết nếu có.",
    },
    {
        "group":    None,
        "stt":      19,
        "hang_muc": "Thêm 3 sản phẩm chuyên biệt mới",
        "mo_ta":    "Thêm OMEGA.RS (Bất động sản), OMEGA.FOT (Retail), OMEGA.DASHBOARD vào trang sản phẩm",
        "files":    "san-pham.html\nsan-pham/omega-rs.html (mới)\nsan-pham/omega-fot.html (mới)\nsan-pham/omega-dashboard.html (mới)",
        "priority": "🔴 Rất cao",
        "loai":     "Nội dung mới",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Cần tạo trang chi tiết cho mỗi sản phẩm + cập nhật sitemap.",
    },
    {
        "group":    None,
        "stt":      20,
        "hang_muc": "Sơ đồ schema sản phẩm — cập nhật text",
        "mo_ta":    "Cập nhật text trong sơ đồ/schema diagram trên trang sản phẩm",
        "files":    "san-pham.html",
        "priority": "🟠 Cao",
        "loai":     "Text + UI",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Nếu dùng SVG inline thì dễ chỉnh. Nếu là ảnh PNG thì cần file mới.",
    },
    {
        "group":    None,
        "stt":      21,
        "hang_muc": "Thêm danh mục Mobile App",
        "mo_ta":    "Thêm category Mobile App với mô tả chi tiết tính năng vào trang sản phẩm",
        "files":    "san-pham.html",
        "priority": "🟠 Cao",
        "loai":     "Nội dung mới",
        "nguon":    "Feedback.docx + BS nội dung.docx",
        "ghi_chu":  "Nội dung mô tả tính năng mobile app đã có trong BS nội dung.docx.",
    },
    {
        "group":    None,
        "stt":      22,
        "hang_muc": "Lợi thế sản phẩm — thêm section",
        "mo_ta":    "Thêm section 5 điểm lợi thế sản phẩm Omega theo BS nội dung.docx",
        "files":    "san-pham.html",
        "priority": "🟡 Trung bình",
        "loai":     "Nội dung mới",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Tăng tính thuyết phục. Có thể đặt sau schema diagram.",
    },
    {
        "group":    None,
        "stt":      23,
        "hang_muc": "Subtext sản phẩm — cập nhật",
        "mo_ta":    "Cập nhật các đoạn text mô tả ngắn (subtext) dưới tên sản phẩm",
        "files":    "san-pham.html",
        "priority": "🟡 Trung bình",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Xem lại từng card sản phẩm.",
    },
    # ══════════════════════ NHÓM 5: KHÁCH HÀNG ══════════════════════
    {
        "group":    "NHÓM 5 — Trang Khách hàng (khach-hang.html)",
        "stt":      24,
        "hang_muc": "Testimonials — 4 người cụ thể",
        "mo_ta":    "Cập nhật testimonials với 4 nhân vật thực: tên + chức vụ + nội dung nhận xét",
        "files":    "khach-hang.html",
        "priority": "🔴 Rất cao",
        "loai":     "Nội dung mới",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Testimonials thực tế tăng conversion rate. Cần ảnh đại diện (nếu có).",
    },
    {
        "group":    None,
        "stt":      25,
        "hang_muc": "Thống kê dịch vụ",
        "mo_ta":    'Cập nhật số liệu: 16+ năm | 98% hài lòng | 1000+ dự án | 24/7 hỗ trợ',
        "files":    "khach-hang.html\nindex.html (nếu có section stats)",
        "priority": "🔴 Rất cao",
        "loai":     "Text",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Con số thực tế — tác động lớn đến trust. Cần đồng bộ ở tất cả nơi hiển thị.",
    },
    {
        "group":    None,
        "stt":      26,
        "hang_muc": "Label/text section Khách hàng",
        "mo_ta":    "Cập nhật các label và text theo yêu cầu feedback trên trang khách hàng",
        "files":    "khach-hang.html",
        "priority": "🟡 Trung bình",
        "loai":     "Text",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "Cần đối chiếu từng mục với file Feedback.docx.",
    },
    {
        "group":    None,
        "stt":      27,
        "hang_muc": "CTA trang Khách hàng",
        "mo_ta":    "Cập nhật nút Call-to-Action trên trang khách hàng theo yêu cầu mới",
        "files":    "khach-hang.html",
        "priority": "🟡 Trung bình",
        "loai":     "Text + UI",
        "nguon":    "Feedback.docx",
        "ghi_chu":  "CTA ảnh hưởng conversion — cần text rõ ràng, hành động cụ thể.",
    },
    # ══════════════════════ NHÓM 6: THIẾT KẾ & GLOBAL ══════════════════════
    {
        "group":    "NHÓM 6 — Thiết kế & Global (toàn bộ site)",
        "stt":      28,
        "hang_muc": "Design spec tổng thể",
        "mo_ta":    "Áp dụng đặc tả thiết kế: nền trắng + gradient xanh + section tối + accent đỏ",
        "files":    "assets/css/omega.css\nindex.html (và tất cả trang có section liên quan)",
        "priority": "🟠 Cao",
        "loai":     "UI/CSS",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Xác nhận scope — áp dụng toàn site hay chỉ trang chủ? Cần review design trước.",
    },
    {
        "group":    None,
        "stt":      29,
        "hang_muc": "Cấu trúc website theo spec mới",
        "mo_ta":    "Đối chiếu cấu trúc site hiện tại với spec trong BS nội dung.docx, cập nhật navigation nếu cần",
        "files":    "auto-omega/gen_sync_header.py\nindex.html (và các trang chứa nav)",
        "priority": "🟠 Cao",
        "loai":     "Cấu trúc",
        "nguon":    "BS nội dung.docx",
        "ghi_chu":  "Kiểm tra: có trang nào cần thêm/xóa khỏi menu? Chạy gen_sync_header.py sau khi thay đổi.",
    },
    # ══════════════════════ NHÓM 7: SITEMAP & SEO ══════════════════════
    {
        "group":    "NHÓM 7 — Sitemap & SEO",
        "stt":      30,
        "hang_muc": "Cập nhật sitemap.xml sau thêm sản phẩm",
        "mo_ta":    "Thêm 3 trang sản phẩm mới (OMEGA.RS, OMEGA.FOT, OMEGA.DASHBOARD) vào sitemap.xml",
        "files":    "sitemap.xml\nauto-omega/gen_sitemap.py",
        "priority": "🟠 Cao",
        "loai":     "SEO",
        "nguon":    "Nội bộ",
        "ghi_chu":  "Chạy gen_sitemap.py sau khi tạo xong các trang mới. Submit lại Google Search Console.",
    },
    {
        "group":    None,
        "stt":      31,
        "hang_muc": "Meta description các trang mới",
        "mo_ta":    "Viết meta description cho 3 trang sản phẩm mới và các trang được tái cấu trúc",
        "files":    "san-pham/omega-rs.html\nsan-pham/omega-fot.html\nsan-pham/omega-dashboard.html",
        "priority": "🟡 Trung bình",
        "loai":     "SEO",
        "nguon":    "Nội bộ",
        "ghi_chu":  "Quan trọng cho SEO — viết đúng format: 120–160 ký tự, chứa keyword.",
    },
]

# ── Tạo Excel ─────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font():  return Font(name="Segoe UI", bold=True, color=C_HEADER_FG, size=11)
def body_font():    return Font(name="Segoe UI", size=10)
def group_font():   return Font(name="Segoe UI", bold=True, size=10, color="1A4731")

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def col_wrap(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def generate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kế hoạch thay đổi"
    ws.freeze_panes = "A3"

    # ── Tiêu đề chính ────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "OMEGA WEBSITE — KẾ HOẠCH CẬP NHẬT & THAY ĐỔI  (Phân tích từ Feedback + BS Nội dung)"
    title_cell.font = Font(name="Segoe UI", bold=True, size=13, color=C_HEADER_FG)
    title_cell.fill = make_fill(C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # ── Header hàng 2 ────────────────────────────────────────────────────────
    headers = ["STT", "Hạng mục thay đổi", "Mô tả chi tiết",
               "File(s) bị tác động", "Độ ưu tiên", "Loại",
               "Nguồn", "Ghi chú / Đánh giá"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font()
        cell.fill = make_fill("1A6B42")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[2].height = 22

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [5, 30, 52, 38, 16, 18, 16, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Ghi dữ liệu ──────────────────────────────────────────────────────────
    row = 3
    for item in CHANGES:
        # Dòng nhóm (group header)
        if item.get("group"):
            ws.merge_cells(f"A{row}:H{row}")
            gc = ws.cell(row=row, column=1, value=item["group"])
            gc.font = group_font()
            gc.fill = make_fill(C_SECTION_BG)
            gc.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1, wrap_text=False)
            gc.border = thin_border()
            ws.row_dimensions[row].height = 18
            row += 1

        # Dòng dữ liệu
        bg = C_ROW_ODD if item["stt"] % 2 == 1 else C_ROW_EVEN
        row_data = [
            item["stt"],
            item["hang_muc"],
            item["mo_ta"],
            item["files"],
            item["priority"],
            item["loai"],
            item["nguon"],
            item["ghi_chu"],
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = body_font()
            cell.border = thin_border()
            cell.alignment = Alignment(
                horizontal="center" if c in (1, 5, 6, 7) else "left",
                vertical="top",
                wrap_text=True,
            )
            # Màu cột ưu tiên
            if c == 5:
                p_color = PRIORITY_COLORS.get(val, bg)
                cell.fill = make_fill(p_color)
                cell.font = Font(name="Segoe UI", size=10,
                                 bold=True, color="FFFFFF" if val in ("🔴 Rất cao",) else "1A1A1A")
            else:
                cell.fill = make_fill(bg)

        ws.row_dimensions[row].height = max(
            30,
            min(90, 15 * max(1, item["files"].count("\n") + 1))
        )
        row += 1

    # ── Sheet 2: Tóm tắt theo nhóm ───────────────────────────────────────────
    ws2 = wb.create_sheet("Tóm tắt & Phân tích")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 50

    summary_data = [
        ("Phân tích tổng hợp", "", ""),
        ("", "", ""),
        ("Nguồn", "Số mục", "Ghi chú"),
        ("Feedback.docx (yêu cầu chỉnh sửa)", 19, "Chủ yếu: text, thương hiệu, UI nhỏ"),
        ("BS nội dung.docx (bổ sung nội dung)", 12, "Chủ yếu: nội dung mới, sản phẩm, thiết kế"),
        ("Nội bộ (kỹ thuật/SEO)", 2, "Sitemap, meta description"),
        ("", "", ""),
        ("Phân loại ưu tiên", "Số mục", ""),
        ("🔴 Rất cao — Cần làm ngay", 9, "Thương hiệu, sản phẩm mới, testimonials"),
        ("🟠 Cao — Làm trong sprint 1", 8, "Trang chủ, sản phẩm, thiết kế global"),
        ("🟡 Trung bình — Sprint 2", 9, "Nội dung phụ, meta, CTA"),
        ("🟢 Thấp — Backlog", 2, "Animation, text nhỏ"),
        ("", "", ""),
        ("Phân loại theo loại công việc", "Số mục", ""),
        ("Text (chỉnh nội dung có sẵn)", 12, "Nhanh, ít rủi ro"),
        ("Nội dung mới (thêm section/module)", 8, "Cần thiết kế + HTML mới"),
        ("UI/CSS (giao diện)", 5, "Cần test cross-browser"),
        ("Cấu trúc (tái tổ chức trang)", 2, "Rủi ro cao — cần review kỹ"),
        ("SEO (sitemap, meta)", 2, "Chạy sau khi xong code"),
        ("", "", ""),
        ("Ước lượng công việc", "", ""),
        ("Tổng số thay đổi", 31, "Trên 7 nhóm trang"),
        ("File HTML bị tác động (ước tính)", "15–20", "Cộng thêm 3 file HTML mới"),
        ("Thứ tự thực hiện đề xuất", "", "Nhóm 1 → 4 → 5 → 2 → 3 → 6 → 7"),
    ]

    for r, (a, b, c) in enumerate(summary_data, 1):
        for ci, val in enumerate([a, b, c], 1):
            cell = ws2.cell(row=r, column=ci, value=val)
            if r == 1:
                cell.font = Font(name="Segoe UI", bold=True, size=13, color=C_HEADER_FG)
                cell.fill = make_fill(C_HEADER_BG)
            elif a in ("Nguồn", "Phân loại ưu tiên", "Phân loại theo loại công việc", "Ước lượng công việc"):
                cell.font = Font(name="Segoe UI", bold=True, size=10, color="1A4731")
                cell.fill = make_fill(C_SECTION_BG)
            else:
                cell.font = Font(name="Segoe UI", size=10)
            cell.alignment = Alignment(
                horizontal="center" if ci == 2 else "left",
                vertical="center", wrap_text=True
            )
        ws2.row_dimensions[r].height = 18
    ws2.merge_cells("A1:C1")

    # ── Sheet 3: Thứ tự thực hiện đề xuất ────────────────────────────────────
    ws3 = wb.create_sheet("Thứ tự thực hiện")
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 55

    ws3.merge_cells("A1:C1")
    h = ws3["A1"]
    h.value = "THỨ TỰ THỰC HIỆN ĐỀ XUẤT"
    h.font = Font(name="Segoe UI", bold=True, size=13, color=C_HEADER_FG)
    h.fill = make_fill(C_HEADER_BG)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    sprint_data = [
        ("Sprint 0\n(Ngay lập tức)", "Đổi tên thương hiệu", "Thay 'OMEGA ERP' → 'OMEGA' / 'Phần mềm của Omega' trên Giải pháp (8 trang) + Dịch vụ + banner. Dùng find-replace có kiểm soát."),
        ("Sprint 1\n(1–3 ngày)", "Sản phẩm & cấu trúc mới", "Tái cấu trúc san-pham.html: nhóm mới + OMEGA.EM + 3 sản phẩm chuyên biệt. Cần design card mới. Tạo 3 trang HTML mới."),
        ("Sprint 1", "Testimonials & stats", "Cập nhật khach-hang.html: 4 testimonials thực + số liệu 98%/1000+. Tác động lớn đến trust."),
        ("Sprint 2\n(3–7 ngày)", "Trang chủ — nội dung", "Tagline, hero subtitle, section headers, ecosystem name, animation background. Cần xác nhận text cuối với khách."),
        ("Sprint 2", "Về Omega — nội dung mới", "Thêm: câu chuyện sáng lập + ý nghĩa logo + 5 giá trị cốt lõi full text + tầm nhìn 2030."),
        ("Sprint 3\n(7–14 ngày)", "Thiết kế & Logo", "Logo gradient xanh (cần file mới), design spec toàn site. Cần phối hợp với designer."),
        ("Sprint 3", "SEO — Sitemap & Meta", "Sau khi code xong: cập nhật sitemap, viết meta description trang mới, submit Google Search Console."),
        ("Backlog\n(sau sprint 3)", "Animation & UX nhỏ", "Animation Về Omega, text nhỏ còn lại. Làm sau khi hoàn thiện nội dung chính."),
    ]

    headers3 = ["Sprint", "Công việc", "Mô tả & lưu ý"]
    for ci, h in enumerate(headers3, 1):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = Font(name="Segoe UI", bold=True, size=10, color=C_HEADER_FG)
        cell.fill = make_fill("1A6B42")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws3.row_dimensions[2].height = 20

    sprint_colors = {
        "Sprint 0": "FF4444",
        "Sprint 1": "FF9900",
        "Sprint 2": "FFD700",
        "Sprint 3": "A8D5A2",
        "Backlog": "D0D0D0",
    }
    for r, (sprint, work, desc) in enumerate(sprint_data, 3):
        color_key = sprint.split("\n")[0]
        bg = sprint_colors.get(color_key, "FFFFFF")
        for ci, val in enumerate([sprint, work, desc], 1):
            cell = ws3.cell(row=r, column=ci, value=val)
            cell.font = Font(name="Segoe UI", size=10, bold=(ci == 1))
            cell.fill = make_fill(bg if ci == 1 else ("F9F9F9" if r % 2 == 1 else "FFFFFF"))
            cell.alignment = Alignment(horizontal="center" if ci == 1 else "left",
                                       vertical="top", wrap_text=True)
            cell.border = thin_border()
        ws3.row_dimensions[r].height = 40

    wb.save(OUTPUT)
    print(f"✅ Đã tạo: {OUTPUT}")
    print(f"   Sheet 1: Kế hoạch thay đổi ({len([x for x in CHANGES if x['stt']])} mục)")
    print(f"   Sheet 2: Tóm tắt & Phân tích")
    print(f"   Sheet 3: Thứ tự thực hiện đề xuất")

if __name__ == "__main__":
    generate()
