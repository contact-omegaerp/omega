"""
_gen_khach_hang_xlsx.py  v2
Tạo khach-hang.xlsx — CMS Khách hàng OMEGA (có sheet ChiTiet)
Tự động trích xuất nội dung từ 19 trang tĩnh khach-hang/*.html
Run: python -X utf8 auto-omega/_gen_khach_hang_xlsx.py
"""

import os, re
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # thư mục gốc project

# ── Màu sắc ──────────────────────────────────────────────────────────────────
def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def hdr_font(color="FFFFFF", size=10, bold=True): return Font(name="Calibri", color=color, size=size, bold=bold)
def data_font(size=10, bold=False, color="1E293B"): return Font(name="Calibri", size=size, bold=bold, color=color)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(wrap=True): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Sheet helpers ─────────────────────────────────────────────────────────────
def write_section_header(ws, row, label, bg, fg="FFFFFF", ncols=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, label)
    c.fill = fill(bg); c.font = Font(name="Calibri", bold=True, size=12, color=fg)
    c.alignment = center(); ws.row_dimensions[row].height = 28

def write_col_headers(ws, row, labels, bg="F1F5F9"):
    for col, lbl in enumerate(labels, 1):
        c = ws.cell(row, col, lbl)
        c.fill = fill(bg); c.font = hdr_font(color="1E293B", size=9)
        c.alignment = center(); c.border = thin_border()
    ws.row_dimensions[row].height = 44

def write_data_row(ws, row, values, even=False):
    bg = "F8FAFC" if even else "FFFFFF"
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        c.fill = fill(bg); c.font = data_font()
        c.alignment = left(); c.border = thin_border()
    ws.row_dimensions[row].height = 18

def write_note(ws, row, text, ncols=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.fill = fill("FFF9C4"); c.font = Font(name="Calibri", italic=True, size=9, color="92400E")
    c.alignment = left(); ws.row_dimensions[row].height = 30

# ── Trích xuất dữ liệu từ trang tĩnh ─────────────────────────────────────────
def extract_detail(slug):
    path = os.path.join(ROOT, "khach-hang", f"{slug}.html")
    if not os.path.exists(path):
        print(f"  WARN: Không có file {slug}.html")
        return {}

    with open(path, encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "html.parser")

    d = {"slug": slug}

    # ── Giới thiệu / Thách thức / Giải pháp ──────────────────────────────
    for card in soup.select(".cp-info-card"):
        lbl_el = card.select_one(".ci-label")
        if not lbl_el: continue
        lbl = lbl_el.get_text(strip=True).lower()
        p_el = card.select_one("p")
        text = p_el.get_text(strip=True) if p_el else ""
        if "giới thiệu" in lbl:   d["gioi_thieu"] = text
        elif "thách thức" in lbl: d["thach_thuc"]  = text
        elif "giải pháp" in lbl:  d["giai_phap"]   = text

    # ── Kết quả (3 result items) ──────────────────────────────────────────
    for i, item in enumerate(soup.select(".cp-result-item")[:3], 1):
        icon_el = item.select_one(".cp-result-icon i")
        num_el  = item.select_one(".cp-result-num")
        lbl_el  = item.select_one(".cp-result-label")
        # FA class: lọc ra lớp fa-xxx (không phải fa-solid)
        icon_cls = ""
        if icon_el:
            classes = [c for c in icon_el.get("class", []) if c.startswith("fa-") and c not in ("fa-solid","fa-regular","fa-brands")]
            icon_cls = classes[0] if classes else "fa-chart-line"
        d[f"ket_qua_{i}_icon"]  = icon_cls
        d[f"ket_qua_{i}_num"]   = num_el.get_text(strip=True)  if num_el  else ""
        d[f"ket_qua_{i}_label"] = lbl_el.get_text(strip=True)  if lbl_el  else ""

    # ── Quote ─────────────────────────────────────────────────────────────
    q_text  = soup.select_one(".cp-q-text")
    q_name  = soup.select_one(".cp-quote-name")
    q_title = soup.select_one(".cp-quote-title")
    # Bỏ dấu nháy kép đầu/cuối nếu có
    qt = q_text.get_text(strip=True) if q_text else ""
    qt = qt.strip('"').strip('"').strip('"').strip("'").strip()
    d["quote_text"]       = qt
    d["quote_nguoi"]      = q_name.get_text(strip=True)  if q_name  else ""
    d["quote_chuc_danh"]  = q_title.get_text(strip=True) if q_title else ""

    # ── Sản phẩm ─────────────────────────────────────────────────────────
    prods = []
    for tag in soup.select(".cp-product-tag"):
        txt = tag.get_text(separator=" ", strip=True)
        # Bỏ icon text nếu có
        txt = re.sub(r"^\s*\S{1,2}\s+", "", txt).strip()
        if txt: prods.append(txt)
    d["san_pham"] = ", ".join(prods)

    # ── H1 / subtitle (page header) cho SEO title ─────────────────────
    h1 = soup.select_one(".page-header h1")
    d["page_h1"] = h1.get_text(strip=True) if h1 else ""

    return d

# ── Dữ liệu KhachHang ────────────────────────────────────────────────────────
KH_HEADERS = [
    "STT","Slug","Tên công ty đầy đủ","Tên ngắn (tooltip)",
    "Tab lọc\n(san-xuat/thuong-mai\n/dich-vu/y-te)","Ngành","Icon FA\n(fa-xxx)",
    "Logo file\n(.png/.jpg)","Hiển thị\n(TRUE/FALSE)","Mô tả tooltip",
    "Câu quote (ngắn)","Người quote",
]
KH_DATA = [
    (1,  "nam-thai-son",    "Công ty CP XNK Nam Thái Sơn",                   "Công ty CP XNK Nam Thái Sơn",    "thuong-mai","Thương mại – XNK",              "fa-boxes-stacking",   "nam-thai-son.png",     True, "Doanh nghiệp xuất nhập khẩu hàng đầu, triển khai OMEGA.ERP toàn diện từ mua hàng, kho vận đến kế toán và báo cáo quản trị.",                                              "ERP không chỉ là phần mềm mà còn là giải pháp quản trị gắn liền với sự đồng hành lâu dài của Omega.", "Ông Vân, Phó Tổng Giám Đốc"),
    (2,  "vipharco",        "Công ty CP Dược phẩm Vipharco",                 "Công ty CP Dược phẩm Vipharco",  "y-te",     "Dược phẩm",                     "fa-capsules",         "vipharco.png",         True, "Triển khai OMEGA.ERP đồng bộ toàn bộ hoạt động dược phẩm: mua hàng, kho, bán hàng, kế toán và báo cáo BI thời gian thực.",                                              "Kiểm kê hàng tồn kho giảm từ 2 ngày xuống chỉ còn 1.5 giờ — nhanh hơn 10 lần.", "Bà Thu, Giám Đốc"),
    (3,  "thanh-nam",       "Tập đoàn Thành Nam",                            "Tập đoàn Thành Nam",             "san-xuat", "Sản xuất – Đa ngành",           "fa-industry",         "thanh-nam.png",        True, "Tập đoàn sản xuất & kinh doanh đa ngành. Sau nhiều lần thử nghiệm ERP, OMEGA.ERP đã giúp chuẩn hóa và đồng bộ toàn bộ hệ thống.",                                          "Sau 3 lần triển khai ERP thất bại, OMEGA.ERP đã vận hành đồng bộ từ sản xuất, kho vận đến kế toán.", "Ông Cường, Tổng Giám Đốc"),
    (4,  "truecare",        "Truecare",                                       "Truecare",                       "y-te",     "Thiết bị y tế",                 "fa-hospital",         "truecare.png",         True, "Phân phối thiết bị y tế chuyên nghiệp. OMEGA.ERP giúp quản lý kho hàng y tế, truy xuất lô số, hạn dùng và đồng bộ kế toán.",                                            "", ""),
    (5,  "cao-su-dau-tieng","Tổng Công ty Cao su Dầu Tiếng",                 "Cao su Dầu Tiếng",               "san-xuat", "Sản xuất cao su",               "fa-leaf",             "cao-su-dau-tieng.png", True, "Doanh nghiệp sản xuất & chế biến cao su thiên nhiên hàng đầu. OMEGA.ERP quản lý quy trình từ khai thác, chế biến đến xuất khẩu.",                                          "", ""),
    (6,  "skypec",          "Công ty Xăng dầu Hàng không Việt Nam (SKYPEC)", "Skypec – Xăng dầu Hàng không",  "thuong-mai","Thương mại – Nhiên liệu",      "fa-oil-can",          "skypec.png",           True, "Công ty Xăng dầu Hàng không Việt Nam ứng dụng OMEGA.ERP để chuẩn hóa quy trình mua – bán – kho và tài chính kế toán toàn hệ thống.",                                      "Giảm 35% thời gian đối soát — tăng độ chính xác kho lên 99%.", ""),
    (7,  "lidovit",         "Lidovit",                                        "Lidovit",                        "y-te",     "Dược phẩm – Thực phẩm chức năng","fa-pills",           "lidovit.png",          True, "Sản xuất và phân phối dược phẩm, thực phẩm chức năng. OMEGA.ERP hỗ trợ quản lý lô số, hạn dùng và hệ thống phân phối đa kênh.",                                           "", ""),
    (8,  "hoa-an",          "Hoa An",                                         "Hoa An",                         "thuong-mai","Thương mại",                   "fa-shop",             "hoa-an.jpg",           True, "Doanh nghiệp thương mại ứng dụng OMEGA.ERP để tối ưu hóa quy trình bán hàng, quản lý công nợ và báo cáo tài chính thời gian thực.",                                        "", ""),
    (9,  "sasco",           "SASCO – Saigon Airport Services Company",        "SASCO – Saigon Airport",         "dich-vu",  "Dịch vụ – Hàng không",          "fa-plane",            "sasco.png",            True, "Dịch vụ thương mại & bán lẻ tại cảng hàng không Tân Sơn Nhất. OMEGA.ERP quản lý hệ thống duty-free, kho vận và tài chính.",                                                "", ""),
    (10, "lyprodan",        "Lyprodan",                                       "Lyprodan",                       "thuong-mai","Phân phối thương mại",         "fa-truck",            "lyprodan.png",         True, "Phân phối hàng hóa thương mại. OMEGA.ERP giúp quản lý hệ thống đại lý, theo dõi đơn hàng và tối ưu dòng tiền toàn chuỗi phân phối.",                                      "", ""),
    (11, "trieu-phu-loc",   "Triều Phú Lộc",                                 "Triều Phú Lộc",                  "san-xuat", "Sản xuất công nghiệp",          "fa-gear",             "trieu-phu-loc.png",    True, "Doanh nghiệp sản xuất công nghiệp triển khai OMEGA.ERP để quản lý lệnh sản xuất, nguyên vật liệu và kiểm soát chi phí hiệu quả.",                                          "", ""),
    (12, "tien-trien",      "Tiến Triển",                                     "Tiến Triển",                     "san-xuat", "Sản xuất",                      "fa-cogs",             "tien-trien.png",       True, "Doanh nghiệp sản xuất và thương mại ứng dụng OMEGA.ERP để số hóa quy trình sản xuất, tối ưu tồn kho và nâng cao hiệu quả.",                                              "", ""),
    (13, "vitajean",        "Vitajean",                                       "Vitajean",                       "san-xuat", "Dệt may – Jeans",               "fa-shirt",            "vitajean.png",         True, "Sản xuất hàng may mặc denim & jeans. OMEGA.ERP giúp quản lý nguyên phụ liệu, quy trình sản xuất từng công đoạn và kiểm soát chất lượng.",                                   "", ""),
    (14, "vstarschool",     "Vstar School",                                   "Vstar School",                   "dich-vu",  "Giáo dục",                      "fa-graduation-cap",   "vstarschool.png",      True, "Hệ thống trường học triển khai OMEGA.EDU & OMEGA.ERP để quản lý học sinh, lịch học, nhân sự giáo viên và tài chính học phí.",                                             "", ""),
    (15, "earth-corp",      "Earth Corp",                                     "Earth Corp",                     "san-xuat", "Sản xuất",                      "fa-industry",         "earth-corp.png",       True, "Tập đoàn sản xuất công nghiệp ứng dụng OMEGA.ERP để quản lý toàn diện chuỗi sản xuất, kho vận và hệ thống báo cáo quản trị.",                                            "", ""),
    (16, "hanel",           "Hanel",                                          "Hanel",                          "thuong-mai","Phân phối điện tử",            "fa-microchip",        "hanel.png",            True, "Tập đoàn phân phối điện tử & thiết bị hàng đầu Hà Nội. OMEGA.ERP giúp quản lý kho điện tử, hệ thống đại lý và tối ưu chuỗi cung ứng.",                                    "", ""),
    (17, "mitsubishi",      "Mitsubishi (đại lý VN)",                         "Mitsubishi",                     "thuong-mai","Phân phối – Thiết bị",         "fa-wrench",           "mitsubishi.png",       True, "Phân phối thiết bị và máy móc Mitsubishi tại Việt Nam. OMEGA.ERP hỗ trợ quản lý đại lý, đơn hàng, kho phụ tùng và dịch vụ bảo hành.",                                     "", ""),
    (18, "him-lam",         "Tập đoàn Him Lam",                              "Him Lam",                        "dich-vu",  "Bất động sản",                  "fa-building-columns", "him-lam.png",          True, "Tập đoàn đầu tư & phát triển bất động sản Him Lam ứng dụng OMEGA.ERP để quản lý dự án, tài chính và nhân sự quy mô lớn.",                                                  "", ""),
    (19, "stdt",            "STDT",                                           "STDT",                           "san-xuat", "Sản xuất công nghiệp",          "fa-gear",             "stdt.png",             True, "Doanh nghiệp sản xuất công nghiệp triển khai OMEGA.ERP để đồng bộ hóa quy trình sản xuất, kiểm soát chi phí và tối ưu năng lực vận hành.",                                "", ""),
]

SLUGS = [row[1] for row in KH_DATA]

# ── Dữ liệu CaseStudy ────────────────────────────────────────────────────────
CS_HEADERS = ["STT","Slug KH","Nhãn ngành","Icon FA","Tiêu đề card","Mô tả","Kết quả nổi bật","Hiển thị"]
CS_DATA = [
    (1,"skypec","Thương mại dầu khí","fa-oil-can","SKYPEC – Triển khai OMEGA.ERP thành công","Công ty Xăng dầu Hàng không Việt Nam ứng dụng OMEGA.ERP để chuẩn hóa quy trình mua – bán – kho và tài chính kế toán trên toàn hệ thống phân phối nhiên liệu.","Giảm 35% thời gian đối soát – Tăng độ chính xác kho 99%",True),
    (2,"vipharco","Dược phẩm","fa-capsules","Vipharco – OMEGA.ERP trong ngành dược phẩm","Công ty Cổ phần Dược phẩm Vipharco triển khai OMEGA.ERP tích hợp quản lý sản xuất, kiểm soát lô – hạn dùng, và hệ thống báo cáo tài chính theo chuẩn ngành dược.","Truy xuất nguồn gốc 100% – Tối ưu tồn kho 28%",True),
    (3,"cao-su-dau-tieng","Sản xuất cao su","fa-industry","Cao su Dầu Tiếng – Triển khai ERP quản lý sản xuất","Tổng Công ty Cao su Dầu Tiếng ứng dụng OMEGA.ERP để số hóa toàn bộ quy trình khai thác – chế biến – bán hàng, tích hợp kho và lương lao động nông trường.","Tiết kiệm 40% nhân lực hành chính – Báo cáo real-time",True),
]

# ── Dữ liệu Testimonial ──────────────────────────────────────────────────────
TM_HEADERS = ["STT","Họ tên","Chức danh","Câu trích dẫn","Số sao","Hiển thị"]
TM_DATA = [
    (1,"Ông Nguyễn Văn Thành","Giám đốc Tài chính – SKYPEC","Sau khi triển khai OMEGA.ERP, chúng tôi kiểm soát được toàn bộ dòng tiền và tồn kho theo thời gian thực. Ban giám đốc có thể xem báo cáo tổng hợp chỉ sau 1 cú click mà không cần chờ bộ phận kế toán tổng hợp như trước.",5,True),
    (2,"Bà Trần Thị Hương","Trưởng phòng Kế hoạch – Vipharco","Đội ngũ tư vấn của Omega hiểu rất rõ đặc thù ngành dược. Từ quản lý lô hàng, hạn sử dụng đến báo cáo theo chuẩn Bộ Y tế – tất cả đều được xử lý trơn tru. Chúng tôi tiết kiệm được đáng kể chi phí vận hành sau 1 năm go-live.",5,True),
    (3,"Ông Lê Quốc Minh","Phó Tổng Giám đốc – Cao su Dầu Tiếng","Hệ thống ERP của Omega giúp chúng tôi quản lý hàng nghìn lao động nông trường và hàng trăm điểm khai thác cao su một cách chính xác. Quy trình chốt lương trước đây mất cả tuần, nay chỉ còn 1 ngày.",4.5,True),
]

# ── Dữ liệu ChiTiet (trích xuất từ trang tĩnh) ──────────────────────────────
CT_HEADERS = [
    "Slug","Tên h1 (SEO)",
    "Giới thiệu","Thách thức","Giải pháp Omega",
    "KQ1 icon","KQ1 số/text","KQ1 mô tả",
    "KQ2 icon","KQ2 số/text","KQ2 mô tả",
    "KQ3 icon","KQ3 số/text","KQ3 mô tả",
    "Quote text","Người quote","Chức danh quote",
    "Sản phẩm (phẩy phân cách)",
]

def build_chi_tiet_data():
    rows = []
    for slug in SLUGS:
        d = extract_detail(slug)
        if not d:
            d = {"slug": slug}
        row = [
            d.get("slug",""),
            d.get("page_h1",""),
            d.get("gioi_thieu",""),
            d.get("thach_thuc",""),
            d.get("giai_phap",""),
            d.get("ket_qua_1_icon","fa-chart-line"),
            d.get("ket_qua_1_num",""),
            d.get("ket_qua_1_label",""),
            d.get("ket_qua_2_icon","fa-users"),
            d.get("ket_qua_2_num",""),
            d.get("ket_qua_2_label",""),
            d.get("ket_qua_3_icon","fa-clock"),
            d.get("ket_qua_3_num",""),
            d.get("ket_qua_3_label",""),
            d.get("quote_text",""),
            d.get("quote_nguoi",""),
            d.get("quote_chuc_danh",""),
            d.get("san_pham",""),
        ]
        rows.append(row)
        print(f"  {slug}: {len(row)} cols, kq={d.get('ket_qua_1_num','?')}, prods={d.get('san_pham','?')[:30]}")
    return rows

# ── Build sheets ──────────────────────────────────────────────────────────────
def build_sheet_kh(wb):
    ws = wb.create_sheet("KhachHang")
    nc = len(KH_HEADERS)
    write_section_header(ws, 1, "📋  KHÁCH HÀNG — Logo Grid", "0D5C38", ncols=nc)
    write_note(ws, 2, "⚠️  Thêm/sửa dòng → GAS tự pickup. URL trang chi tiết tự động tính theo slug (khach-hang/index.html?slug=xxx).", ncols=nc)
    write_col_headers(ws, 3, KH_HEADERS)
    for i, row in enumerate(KH_DATA, 4):
        write_data_row(ws, i, row, even=(i%2==0))
    widths = [5,18,38,32,14,26,18,20,10,55,55,22]
    for ci, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"

def build_sheet_cs(wb):
    ws = wb.create_sheet("CaseStudy")
    nc = len(CS_HEADERS)
    write_section_header(ws, 1, "🏆  CASE STUDY — Featured cards", "00A651", ncols=nc)
    write_note(ws, 2, "⚠️  Tối đa 3 dòng hien_thi=TRUE. Thứ tự theo cột STT.", ncols=nc)
    write_col_headers(ws, 3, CS_HEADERS)
    for i, row in enumerate(CS_DATA, 4):
        write_data_row(ws, i, row, even=(i%2==0))
    widths = [5,18,22,14,42,65,50,10]
    for ci, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"

def build_sheet_tm(wb):
    ws = wb.create_sheet("Testimonial")
    nc = len(TM_HEADERS)
    write_section_header(ws, 1, "💬  ĐÁNH GIÁ — Testimonial Swiper", "0D1B2A", ncols=nc)
    write_note(ws, 2, "⚠️  so_sao: 5 hoặc 4.5.", ncols=nc)
    write_col_headers(ws, 3, TM_HEADERS)
    for i, row in enumerate(TM_DATA, 4):
        write_data_row(ws, i, row, even=(i%2==0))
    widths = [5,26,34,80,12,10]
    for ci, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"

def build_sheet_ct(wb, ct_rows):
    ws = wb.create_sheet("ChiTiet")
    nc = len(CT_HEADERS)
    write_section_header(ws, 1, "📄  CHI TIẾT KHÁCH HÀNG — Trang khach-hang/index.html?slug=xxx", "1E3A5F", ncols=nc)
    write_note(ws, 2, "⚠️  Mỗi dòng = 1 khách hàng. Cột 'Slug' phải khớp với cột Slug ở sheet KhachHang.", ncols=nc)
    write_col_headers(ws, 3, CT_HEADERS)
    for i, row in enumerate(ct_rows, 4):
        write_data_row(ws, i, row, even=(i%2==0))
        ws.row_dimensions[i].height = 22
    widths = [18,30,55,55,55,14,14,34,14,14,34,14,14,34,65,26,30,50]
    for ci, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "B4"

def build_sheet_guide(wb):
    """Sheet hướng dẫn cho nhân viên quản trị web."""
    ws = wb.create_sheet("Guide")
    ws.sheet_view.showGridLines = False

    NC = 4  # 4 cột: A=nhãn, B=nội dung, C=ví dụ, D=ghi chú

    def cw(col, w): ws.column_dimensions[get_column_letter(col)].width = w
    cw(1, 28); cw(2, 62); cw(3, 40); cw(4, 30)

    # ── Helper cells ──────────────────────────────────────────────────────────
    def rh(r, h): ws.row_dimensions[r].height = h

    def title_row(r, txt):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(r, 1, txt)
        c.fill = fill("0D5C38"); c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(r, 32)

    def section(r, txt, bg="0D1B2A"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(r, 1, txt)
        c.fill = fill(bg); c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(r, 26)

    def sub(r, txt, bg="E2E8F0"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(r, 1, txt)
        c.fill = fill(bg); c.font = Font(name="Calibri", bold=True, size=10, color="1E293B")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(r, 22)

    def th(r, labels):  # table column header
        bgs = ["1E293B"] * len(labels)
        for ci, (lbl, bg) in enumerate(zip(labels, bgs), 1):
            c = ws.cell(r, ci, lbl)
            c.fill = fill("1E293B"); c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
        rh(r, 36)

    def td(r, vals, even=False, hl=None):
        bg = "F8FAFC" if even else "FFFFFF"
        if hl: bg = hl
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.fill = fill(bg)
            c.font = Font(name="Calibri", size=9, color="1E293B")
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = thin_border()
        rh(r, 20 if not any(len(str(v or "")) > 50 for v in vals) else 32)

    def note(r, txt, bg="FFF9C4", fg="92400E"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(r, 1, txt)
        c.fill = fill(bg); c.font = Font(name="Calibri", italic=True, size=9, color=fg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        rh(r, 28)

    def blank(r, h=8):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        rh(r, h)

    def step(r, num, txt, example=""):
        ws.cell(r, 1, f"Bước {num}").font = Font(name="Calibri", bold=True, size=10, color="00A651")
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(r, 2, txt).font   = Font(name="Calibri", size=10, color="1E293B")
        ws.cell(r, 2).alignment   = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if example:
            ws.cell(r, 4, example).font      = Font(name="Calibri", italic=True, size=9, color="2563EB")
            ws.cell(r, 4).alignment           = Alignment(horizontal="left", vertical="top", wrap_text=True)
        rh(r, 22)

    # ════════════════════════════════════════════════════════
    r = 1
    title_row(r, "📋  HƯỚNG DẪN CMS KHÁCH HÀNG — Dành cho nhân viên quản trị website")
    r = 2; blank(r, 10)

    # ── SECTION 1: SƠ ĐỒ HỆ THỐNG ──────────────────────────
    r = 3; section(r, "① SƠ ĐỒ HỆ THỐNG — Cách hoạt động")
    r = 4; sub(r, "Luồng dữ liệu từ Sheet đến Website")
    r = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1,
        "Google Sheet  →  Google Apps Script (GAS)  →  Website omega.com.vn\n"
        "Nhân viên sửa Sheet  →  Xóa cache (menu 📋 OMEGA CMS → Xóa cache)  →  Trang web tự cập nhật trong < 1 phút")
    c.font = Font(name="Calibri", size=10, color="1E293B")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    c.fill = fill("EFF6FF")
    rh(r, 44)

    r = 6; th(r, ["Sheet", "Dùng để làm gì", "Cập nhật khi nào", "Key chính"])
    for i, row in enumerate([
        ("KhachHang",  "Logo grid – danh sách tất cả khách hàng",           "Thêm/ẩn/sửa thông tin logo",         "Slug"),
        ("CaseStudy",  "3 thẻ 'Case Study nổi bật' trên trang chính",        "Thay đổi case study được featured",   "Slug KH"),
        ("Testimonial","Carousel đánh giá khách hàng trên trang chính",      "Thêm/sửa câu trích dẫn, số sao",     "—"),
        ("ChiTiet",    "Nội dung trang chi tiết từng khách hàng",            "Cập nhật nội dung hồ sơ khách hàng", "Slug"),
    ], start=1):
        td(r + i, row, even=(i%2==0))
    r = 11

    r += 1; note(r, "🔑  KEY CHÍNH: Cột 'Slug' liên kết KhachHang ↔ ChiTiet. Slug phải KHỚP chính xác (phân biệt hoa/thường, không khoảng trắng).")
    r += 1; blank(r, 10)

    # ── SECTION 2: QUY TẮC SLUG ──────────────────────────────
    r += 1; section(r, "② QUY TẮC SLUG — Khóa chính định danh khách hàng")
    r += 1; th(r, ["Quy tắc", "Giải thích", "Ví dụ đúng ✅", "Ví dụ sai ❌"])
    rules = [
        ("Chỉ dùng chữ thường",      "Không dùng chữ HOA",                          "skypec",             "SKYPEC"),
        ("Không dấu tiếng Việt",      "Chuyển đổi: ă→a, ơ→o, đ→d, ệ→e...",          "cao-su-dau-tieng",   "cao-su-dầu-tiếng"),
        ("Dùng dấu gạch ngang",       "Thay khoảng trắng bằng - (dash)",             "nam-thai-son",       "nam thai son"),
        ("Không ký tự đặc biệt",     "Không !, @, #, /, \\, &...",                   "him-lam",            "him&lam"),
        ("Ngắn gọn, có nghĩa",        "Không cần ghi đầy đủ tên công ty",            "vipharco",           "cong-ty-cp-duoc-pham-vipharco"),
        ("Slug = tên file ảnh",       "File ảnh phải trùng tên với slug",            "skypec.webp",        "xang-dau-hk.webp"),
    ]
    for i, row in enumerate(rules, 1):
        td(r + i, row, even=(i%2==0))
    r += len(rules) + 1

    r += 1; note(r, "💡  Công cụ chuyển slug: dùng https://slugify.online hoặc nhờ kỹ thuật tạo hộ. Sau khi đã dùng, KHÔNG đổi slug vì sẽ làm mất link cũ.")
    r += 1; blank(r, 10)

    # ── SECTION 3: HÌNH ẢNH ──────────────────────────────────
    r += 1; section(r, "③ HÌNH ẢNH LOGO — Cách chuẩn bị và upload")
    r += 1; th(r, ["Tiêu chí", "Yêu cầu", "Lý do", ""])
    img_rules = [
        ("Định dạng ưu tiên",  "WebP (.webp)",                   "Nhỏ hơn PNG 30-40%, tải nhanh hơn",       ""),
        ("Định dạng thay thế", "PNG (.png) hoặc JPG (.jpg)",     "Dùng nếu không xuất được WebP",            ""),
        ("Kích thước",         "300 × 200 px (tỷ lệ 3:2)",       "Vừa với khung logo box 150×80px trên web", ""),
        ("Nền logo",           "Trong suốt (transparent)",        "WebP/PNG hỗ trợ nền trong, JPG thì không", ""),
        ("Dung lượng",         "< 100 KB / ảnh",                  "Tối ưu tốc độ tải trang",                  ""),
        ("Tên file",           "Giống slug + đuôi file",          "Hệ thống tự ghép đường dẫn",              "skypec.webp"),
        ("Thư mục upload",     "khach-hang/  (trong Git repo)",   "GAS đọc logo từ đường dẫn này",           "khach-hang/skypec.webp"),
        ("Cột Logo file",      "Chỉ ghi tên file, không ghi đường dẫn đầy đủ", "Hệ thống tự thêm prefix",  "skypec.webp  (không phải khach-hang/skypec.webp)"),
    ]
    for i, row in enumerate(img_rules, 1):
        td(r + i, row, even=(i%2==0))
    r += len(img_rules) + 1

    r += 1; sub(r, "Quy trình upload ảnh lên Git", "D6EAF8")
    r += 1; step(r, 1, "Chuẩn bị ảnh logo: xuất WebP 300×200px, nền trong suốt nếu có thể", "Dùng Figma, Photoshop, hoặc squoosh.app")
    r += 1; step(r, 2, "Đặt tên file: giống slug, ví dụ slug='vipharco' → file='vipharco.webp'", "vipharco.webp")
    r += 1; step(r, 3, "Vào thư mục khach-hang/ trong Git repo, kéo thả file vào (hoặc commit)", "GitHub Desktop hoặc VS Code Source Control")
    r += 1; step(r, 4, "Trong sheet KhachHang, cột 'Logo file': ghi tên file (ví dụ: vipharco.webp)", "vipharco.webp")
    r += 1; note(r, "⚠️  Nếu ảnh không hiển thị: kiểm tra (1) tên file có khớp slug không, (2) file đã được push lên Git repo chưa.")
    r += 1; blank(r, 10)

    # ── SECTION 4: THÊM KHÁCH HÀNG MỚI ─────────────────────
    r += 1; section(r, "④ THÊM KHÁCH HÀNG MỚI — Quy trình đầy đủ")
    r += 1; step(r, 1, "Chuẩn bị slug: chuyển tên công ty sang dạng slug (vd: 'Công ty ABC' → 'abc')", "'abc'")
    r += 1; step(r, 2, "Upload ảnh logo vào thư mục khach-hang/ trên Git (đặt tên: abc.webp)", "abc.webp")
    r += 1; step(r, 3, "Sheet KhachHang: thêm 1 dòng mới cuối, điền đầy đủ tất cả cột, đặt Hiển thị = TRUE", "Xem bảng độ dài ở Mục ⑤")
    r += 1; step(r, 4, "Sheet ChiTiet: thêm 1 dòng mới, Slug phải KHỚP với slug ở Bước 1", "Slug phân biệt hoa/thường: abc ≠ ABC")
    r += 1; step(r, 5, "(Tùy chọn) Sheet CaseStudy: thêm dòng nếu muốn featured, đặt Hiển thị = TRUE", "Tối đa 3 dòng featured")
    r += 1; step(r, 6, "Xóa cache: menu '📋 OMEGA CMS' → 'Xóa toàn bộ cache'", "Hoặc chờ 5 phút cache tự hết hạn")
    r += 1; step(r, 7, "Kiểm tra: mở trang khach-hang.html và khach-hang/index.html?slug=abc", "Trang chi tiết = ?slug=abc")
    r += 1; note(r, "🙈  ẨN khách hàng (không xóa): đổi cột Hiển thị = FALSE. Dữ liệu vẫn còn trong Sheet, có thể bật lại bất cứ lúc nào.")
    r += 1; blank(r, 10)

    # ── SECTION 5: ĐỘ DÀI VĂN BẢN ──────────────────────────
    r += 1; section(r, "⑤ ĐỘ DÀI VĂN BẢN — Khuyến nghị cho từng trường")
    r += 1; th(r, ["Trường", "Sheet", "Khoảng ký tự", "Ghi chú & Ví dụ"])
    lengths = [
        ("Slug",                   "KhachHang", "5 – 25",    "Chỉ a-z, 0-9, dấu gạch ngang. Vd: 'cao-su-dau-tieng'"),
        ("Tên công ty đầy đủ",     "KhachHang", "30 – 80",   "Tên chính thức đầy đủ. Vd: 'Công ty CP Dược phẩm Vipharco'"),
        ("Tên ngắn (tooltip)",     "KhachHang", "15 – 35",   "Hiển thị tiêu đề tooltip. Vd: 'Vipharco'"),
        ("Ngành nghề",             "KhachHang", "10 – 30",   "Vd: 'Dược phẩm', 'Thương mại – XNK'"),
        ("Mô tả tooltip",          "KhachHang", "100 – 160", "Đoạn ngắn khi hover logo. Không cần quá chi tiết."),
        ("Quote ngắn (tooltip)",   "KhachHang", "60 – 100",  "Trích dẫn ngắn trong popup. Vd: 'Giảm 35% thời gian đối soát'"),
        ("Người quote (tooltip)",  "KhachHang", "15 – 35",   "Vd: 'Ông Nguyễn Văn A, TGĐ'"),
        ("—",                      "—",         "—",         ""),
        ("Tên H1 (SEO)",           "ChiTiet",   "40 – 80",   "Tiêu đề trang chi tiết. Vd: 'Công ty Xăng dầu HK Việt Nam (SKYPEC)'"),
        ("Giới thiệu",             "ChiTiet",   "150 – 300", "Giới thiệu công ty, lĩnh vực hoạt động, quy mô"),
        ("Thách thức",             "ChiTiet",   "150 – 280", "Vấn đề doanh nghiệp gặp trước khi dùng ERP"),
        ("Giải pháp Omega",        "ChiTiet",   "150 – 280", "Cách OMEGA.ERP giải quyết vấn đề của họ"),
        ("KQ1/2/3 — Số/text",      "ChiTiet",   "3 – 15",    "Con số nổi bật. Vd: '35%', '1 click', 'Real-time'"),
        ("KQ1/2/3 — Mô tả",        "ChiTiet",   "25 – 50",   "Giải thích con số. Vd: 'Giảm thời gian đối soát chứng từ'"),
        ("Quote text (đầy đủ)",    "ChiTiet",   "100 – 250", "Trích dẫn đầy đủ từ lãnh đạo, để trong dấu ngoặc kép"),
        ("Người quote",            "ChiTiet",   "15 – 30",   "Vd: 'Ông Nguyễn Văn Thành'"),
        ("Chức danh quote",        "ChiTiet",   "20 – 45",   "Vd: 'Giám đốc Tài chính – SKYPEC'"),
        ("Sản phẩm",               "ChiTiet",   "Phẩy phân cách", "Vd: 'OMEGA.ERP, OMEGA.GL, OMEGA.HR'"),
        ("—",                      "—",         "—",         ""),
        ("Nhãn ngành (Case Study)","CaseStudy", "10 – 25",   "Vd: 'Thương mại dầu khí', 'Dược phẩm'"),
        ("Tiêu đề Case Study",     "CaseStudy", "40 – 75",   "Vd: 'SKYPEC – Triển khai OMEGA.ERP thành công'"),
        ("Mô tả Case Study",       "CaseStudy", "100 – 200", "Mô tả ngắn về dự án triển khai"),
        ("Kết quả Case Study",     "CaseStudy", "50 – 80",   "Số liệu nổi bật. Vd: 'Giảm 35% đối soát – Kho chính xác 99%'"),
    ]
    for i, row in enumerate(lengths, 1):
        hl = "F1F5F9" if row[0] == "—" else ("F8FAFC" if i % 2 == 0 else "FFFFFF")
        td(r + i, row, hl=hl)
    r += len(lengths) + 1

    r += 1; note(r, "📏  Độ dài là khuyến nghị, không phải giới hạn cứng. Quá ngắn → thiếu thông tin. Quá dài → bị cắt xén hoặc khó đọc trên màn hình nhỏ.")
    r += 1; blank(r, 10)

    # ── SECTION 6: DANH SÁCH SẢN PHẨM OMEGA ────────────────
    r += 1; section(r, "⑥ DANH SÁCH SẢN PHẨM OMEGA — Dùng trong cột 'Sản phẩm' sheet ChiTiet")
    r += 1; sub(r, "Phẩy phân cách các sản phẩm, ví dụ: OMEGA.ERP, OMEGA.GL, OMEGA.HR, OMEGA.PR")
    r += 1; th(r, ["Mã sản phẩm", "Tên đầy đủ", "Nhóm", ""])
    products = [
        ("OMEGA.ERP",       "Phần mềm ERP tích hợp",                   "Nền tảng", ""),
        ("OMEGA.SM",        "Quản trị hệ thống",                        "Nền tảng", ""),
        ("OMEGA.SD",        "Thông tin dùng chung",                     "Nền tảng", ""),
        ("OMEGA.GL",        "Kế toán doanh nghiệp",                     "Tài chính", ""),
        ("OMEGA.FA",        "Kế toán tài sản cố định",                  "Tài chính", ""),
        ("OMEGA.MC",        "Kế toán quản trị",                         "Tài chính", ""),
        ("OMEGA.CL",        "Báo cáo tài chính hợp nhất",               "Tài chính", ""),
        ("OMEGA.HR",        "Quản lý nhân sự",                          "Nhân sự",   ""),
        ("OMEGA.PR",        "Quản lý tiền lương",                       "Nhân sự",   ""),
        ("OMEGA.MM",        "Quản lý sản xuất",                         "Sản xuất",  ""),
        ("OMEGA.PC",        "Tính giá thành sản phẩm",                  "Sản xuất",  ""),
        ("OMEGA.QC",        "Quản lý chất lượng",                       "Sản xuất",  ""),
        ("OMEGA.EM",        "Quản lý bảo trì bảo dưỡng",                "Sản xuất",  ""),
        ("OMEGA.PO",        "Quản lý mua hàng",                         "Thương mại",""),
        ("OMEGA.SO",        "Quản lý bán hàng",                         "Thương mại",""),
        ("OMEGA.WM",        "Quản lý kho/tồn kho",                      "Thương mại",""),
        ("GAMA.SMB",        "Kế toán chuyên nghiệp cho DN vừa & nhỏ",  "Standalone",""),
        ("OMEGA.CRM",       "Quản lý quan hệ khách hàng",               "Standalone",""),
        ("OMEGA.EDU",       "Quản lý đào tạo & trường học",             "Standalone",""),
        ("OMEGA.RS",        "Quản lý BĐS cho thuê",                     "Standalone",""),
        ("OMEGA.FOT",       "Quản lý bán lẻ (chuỗi cửa hàng)",         "Standalone",""),
        ("OMEGA.DASHBOARD", "Phân tích tài chính & BI",                 "Standalone",""),
        ("OMEGA.APV",       "App duyệt chứng từ (iOS/Android)",         "Mobile App",""),
        ("OMEGA.SCR",       "App quét mã vạch kho",                     "Mobile App",""),
        ("OMEGA.STK",       "App tra cứu tồn kho",                      "Mobile App",""),
        ("OMEGA.SOR",       "App quản lý đặt hàng",                     "Mobile App",""),
        ("OMEGA.MST",       "App thống kê sản xuất",                    "Mobile App",""),
        ("OMEGA.HRM",       "App quản trị nhân lực",                    "Mobile App",""),
    ]
    grp_colors = {"Nền tảng":"EDE9FE","Tài chính":"DCFCE7","Nhân sự":"DBEAFE","Sản xuất":"FEF3C7","Thương mại":"FCE7F3","Standalone":"F1F5F9","Mobile App":"E0F2FE"}
    for i, row in enumerate(products, 1):
        hl = grp_colors.get(row[2], "FFFFFF")
        td(r + i, row, hl=hl)
    r += len(products) + 1

    r += 1; blank(r, 10)

    # ── SECTION 7: ICON FA & DROPDOWN ───────────────────────
    r += 1; section(r, "⑦ ICON FONT AWESOME — Tham chiếu, dropdown ở sheet Icons")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Danh sách icon đầy đủ nằm trong sheet 'Icons'. Để thiết lập dropdown trong Google Sheets:\n"
        "Data → Data Validation → Criteria: 'List from a range' → chọn cột A của sheet Icons (Icons!A:A)")
    c.font = Font(name="Calibri", size=10, color="1E293B")
    c.fill = fill("EFF6FF")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    rh(r, 44)

    r += 1; blank(r, 10)

    # ── SECTION 8: CACHE & CẬP NHẬT ─────────────────────────
    r += 1; section(r, "⑧ CACHE & CẬP NHẬT NHANH")
    r += 1; th(r, ["Tình huống", "Thời gian cập nhật", "Cách force update ngay", ""])
    cache_rows = [
        ("Sửa nội dung bất kỳ trong Sheet",    "≤ 5 phút (cache GAS tự hết hạn)",   "Menu '📋 OMEGA CMS' → 'Xóa toàn bộ cache'",  ""),
        ("Thêm khách hàng mới",                "≤ 5 phút",                           "Xóa cache sau khi thêm xong",                 ""),
        ("Upload ảnh mới lên Git",             "Tức thì sau khi push",               "Hard refresh trình duyệt (Ctrl+Shift+R)",      ""),
        ("Người dùng duyệt lại trang cũ",      "≤ 30 phút (cache localStorage)",     "Người dùng Ctrl+Shift+R hoặc xóa cache trình duyệt", ""),
        ("Deploy GAS mới",                     "Cần cập nhật URL trong 2 file HTML", "Xem bước triển khai trong sheet HuongDan",    ""),
    ]
    for i, row in enumerate(cache_rows, 1):
        td(r + i, row, even=(i%2==0))
    r += len(cache_rows) + 2

    r += 1; note(r, "✅  Quy trình chuẩn: Sửa Sheet → Ctrl+S → Menu '📋 OMEGA CMS → Xóa cache' → Tải lại trang web để kiểm tra.")


def build_sheet_icons(wb):
    """Sheet danh sách icon FA — dùng làm nguồn cho dropdown validation."""
    ws = wb.create_sheet("Icons")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45

    def rh(r, h): ws.row_dimensions[r].height = h

    # Header
    ws.merge_cells("A1:C1")
    c = ws.cell(1, 1, "📌  DANH SÁCH ICON FONT AWESOME 6 — Dùng cho cột 'Icon FA' trong sheet KhachHang & CaseStudy")
    c.fill = fill("0D5C38"); c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(1, 26)

    ws.merge_cells("A2:C2")
    c = ws.cell(2, 1, "Cách dùng: Data → Data Validation → List from a range → chọn Icons!A3:A999  •  Chỉ nhập đúng mã icon (cột A), không nhập mô tả")
    c.fill = fill("FFF9C4"); c.font = Font(name="Calibri", italic=True, size=9, color="92400E")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(2, 22)

    for ci, lbl in enumerate(["Mã icon (dùng trong Sheet)", "Tên hiển thị", "Thường dùng cho ngành"], 1):
        c = ws.cell(3, ci, lbl)
        c.fill = fill("1E293B"); c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    rh(3, 30)

    icons = [
        # Sản xuất
        ("fa-industry",          "Nhà máy công nghiệp",      "Sản xuất công nghiệp, chế biến"),
        ("fa-factory",           "Xưởng sản xuất",           "Nhà máy, khu công nghiệp"),
        ("fa-gear",              "Bánh răng",                "Cơ khí, chế tạo, sản xuất"),
        ("fa-cogs",              "Nhiều bánh răng",          "Máy móc tự động hóa"),
        ("fa-screwdriver-wrench","Công cụ SX",               "Sản xuất cơ khí, lắp ráp"),
        # Kho vận & Phân phối
        ("fa-boxes-stacking",    "Thùng hàng chồng",         "Kho hàng, XNK, logistics"),
        ("fa-warehouse",         "Nhà kho",                  "Kho vận, phân phối"),
        ("fa-truck",             "Xe tải",                   "Vận chuyển, phân phối"),
        ("fa-truck-fast",        "Xe tải nhanh",             "Giao hàng nhanh, express"),
        ("fa-ship",              "Tàu thủy",                 "Vận tải biển, XNK"),
        ("fa-plane",             "Máy bay",                  "Hàng không, dịch vụ sân bay"),
        # Năng lượng & Tài nguyên
        ("fa-oil-can",           "Bình dầu",                 "Xăng dầu, nhiên liệu"),
        ("fa-gas-pump",          "Cột xăng",                 "Trạm xăng, nhiên liệu"),
        ("fa-bolt",              "Tia sét",                  "Điện, năng lượng"),
        ("fa-solar-panel",       "Pin mặt trời",             "Năng lượng tái tạo"),
        # Nông nghiệp & Thực phẩm
        ("fa-leaf",              "Lá cây",                   "Cao su, nông nghiệp, môi trường"),
        ("fa-seedling",          "Cây non",                  "Nông nghiệp, trồng trọt"),
        ("fa-fish",              "Con cá",                   "Thủy hải sản, nuôi trồng"),
        ("fa-cow",               "Con bò",                   "Chăn nuôi, sữa, thực phẩm"),
        ("fa-utensils",          "Dao dĩa",                  "F&B, nhà hàng, thực phẩm"),
        ("fa-burger",            "Hamburger",                "Fast food, chuỗi F&B"),
        ("fa-bottle-water",      "Chai nước",                "Nước uống, đồ uống"),
        # Y tế & Dược phẩm
        ("fa-capsules",          "Viên nang thuốc",          "Dược phẩm, sản xuất thuốc"),
        ("fa-pills",             "Viên thuốc",               "Dược phẩm, TPCN"),
        ("fa-hospital",          "Bệnh viện",                "Bệnh viện, thiết bị y tế"),
        ("fa-stethoscope",       "Ống nghe tim",             "Y tế, phòng khám"),
        ("fa-microscope",        "Kính hiển vi",             "Lab, xét nghiệm, R&D"),
        ("fa-flask",             "Bình thí nghiệm",          "Hóa chất, nghiên cứu"),
        # Dệt may & Thời trang
        ("fa-shirt",             "Áo sơ mi",                 "Dệt may, thời trang"),
        ("fa-scissors",          "Kéo",                      "May mặc, thủ công"),
        ("fa-gem",               "Viên đá quý",              "Trang sức, hàng cao cấp"),
        ("fa-shoe-prints",       "Dấu giày",                 "Giày dép, thời trang"),
        # Công nghệ & Điện tử
        ("fa-microchip",         "Vi mạch",                  "Điện tử, công nghệ, IT"),
        ("fa-computer",          "Máy tính",                 "IT, phần mềm, dịch vụ số"),
        ("fa-network-wired",     "Mạng dây",                 "Viễn thông, hạ tầng mạng"),
        ("fa-satellite-dish",    "Ăng-ten chảo",             "Viễn thông, truyền thông"),
        # BĐS & Tài chính
        ("fa-building-columns",  "Tòa nhà cổ điển",         "BĐS, ngân hàng, tài chính"),
        ("fa-building",          "Tòa nhà",                  "Bất động sản, văn phòng"),
        ("fa-landmark",          "Công trình lịch sử",       "Tập đoàn lớn, tổ chức"),
        ("fa-coins",             "Đồng tiền",                "Tài chính, đầu tư"),
        ("fa-chart-line",        "Biểu đồ tăng",             "Chứng khoán, phân tích TC"),
        ("fa-sack-dollar",       "Túi tiền",                 "Quỹ đầu tư, tài chính"),
        ("fa-shield-halved",     "Tấm khiên nửa",            "Bảo hiểm, an ninh, bảo vệ"),
        # Xây dựng & Vật liệu
        ("fa-hammer",            "Búa",                      "Xây dựng, vật liệu XD"),
        ("fa-hard-hat",          "Mũ bảo hộ",                "Xây dựng, an toàn lao động"),
        ("fa-paint-roller",      "Con lăn sơn",              "Nội thất, trang trí, sơn"),
        ("fa-wrench",            "Cờ lê",                    "Cơ khí, thiết bị, bảo trì"),
        # Bán lẻ & Thương mại
        ("fa-shop",              "Cửa hàng",                 "Bán lẻ, thương mại nhỏ"),
        ("fa-store",             "Siêu thị nhỏ",             "Chuỗi cửa hàng, retail"),
        ("fa-basket-shopping",   "Giỏ hàng",                 "TMĐT, mua sắm online"),
        # Giáo dục
        ("fa-graduation-cap",    "Mũ tốt nghiệp",            "Giáo dục, đào tạo, trường học"),
        ("fa-book-open",         "Sách mở",                  "Xuất bản, giáo dục"),
        ("fa-chalkboard-user",   "Bảng giảng bài",           "Trung tâm đào tạo"),
        # Du lịch & Dịch vụ
        ("fa-hotel",             "Khách sạn",                "Du lịch, nghỉ dưỡng"),
        ("fa-spa",               "Hoa sen",                  "Spa, làm đẹp, chăm sóc"),
        # Vận tải
        ("fa-car",               "Ô tô",                     "Ô tô, xe cộ, phân phối xe"),
        # In ấn & Bao bì
        ("fa-print",             "Máy in",                   "In ấn, bao bì, xuất bản"),
        ("fa-box",               "Hộp carton",               "Bao bì, đóng gói"),
        # Môi trường
        ("fa-recycle",           "Mũi tên tái chế",          "Môi trường, tái chế"),
        # Khác
        ("fa-globe",             "Quả địa cầu",              "Xuất khẩu, quốc tế, đa quốc gia"),
        ("fa-hand-holding-heart","Lòng bàn tay trái tim",   "Phi lợi nhuận, CSR, dịch vụ xã hội"),
        ("fa-briefcase",         "Cặp công sở",              "Dịch vụ chuyên nghiệp, tư vấn"),
    ]

    grp_ranges = {
        "Sản xuất":       (0, 5,  "FEF3C7"),
        "Kho vận":        (5, 11, "DCFCE7"),
        "Năng lượng":     (11, 15,"FDE8D8"),
        "Nông nghiệp":    (15, 22,"D1FAE5"),
        "Y tế":           (22, 28,"DBEAFE"),
        "Dệt may":        (28, 32,"FCE7F3"),
        "Công nghệ":      (32, 36,"EDE9FE"),
        "BĐS&TC":         (36, 43,"FEF9C3"),
        "Xây dựng":       (43, 47,"F1F5F9"),
        "Bán lẻ":         (47, 50,"FFE4E6"),
        "Giáo dục":       (50, 53,"E0F2FE"),
        "Dịch vụ":        (53, 55,"F5F5F5"),
        "Khác":           (55, 99,"FAFAFA"),
    }
    color_map = {}
    for grp, (s, e, col) in grp_ranges.items():
        for i in range(s, e):
            color_map[i] = col

    for i, (code, name, desc) in enumerate(icons):
        row_idx = i + 4
        bg = color_map.get(i, "FFFFFF")
        for ci, val in enumerate([code, name, desc], 1):
            c = ws.cell(row_idx, ci, val)
            c.fill = fill(bg)
            c.font = Font(name="Calibri", size=9,
                          bold=(ci==1), color="1E293B")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border()
        rh(row_idx, 18)

    ws.freeze_panes = "A4"


def build_sheet_hd(wb):
    ws = wb.create_sheet("HuongDan")
    ws.column_dimensions["A"].width = 110
    lines = [
        ("🚀 HƯỚNG DẪN CMS KHÁCH HÀNG OMEGA v2", "0D5C38"),
        ("", None),
        ("BƯỚC 1 — Import file này lên Google Sheets:", "1E293B"),
        ("   File → Import → Upload → chọn khach-hang.xlsx → Import data → Replace spreadsheet", None),
        ("", None),
        ("BƯỚC 2 — Gắn Google Apps Script:", "1E293B"),
        ("   Extensions → Apps Script → dán nội dung khach-hang.gs → Ctrl+S", None),
        ("", None),
        ("BƯỚC 3 — Deploy Web App:", "1E293B"),
        ("   Deploy → New deployment → Web app → Execute as: Me | Who has access: Anyone → Deploy", None),
        ("   Copy URL deploy (dạng https://script.google.com/macros/s/.../exec)", None),
        ("", None),
        ("BƯỚC 4 — Gắn URL vào trang web:", "1E293B"),
        ("   Mở khach-hang.html VÀ khach-hang/index.html → tìm 'PASTE_YOUR_GAS_WEB_APP_URL_HERE' → thay URL", None),
        ("", None),
        ("BƯỚC 5 — Xóa 19 trang tĩnh (tùy chọn):", "1E293B"),
        ("   Xóa tất cả file trong thư mục khach-hang/ (trừ index.html mới tạo)", None),
        ("   Hoặc giữ lại để backward-compatible với link cũ", None),
        ("", None),
        ("LƯU Ý:", "00A651"),
        ("   · Cột 'Slug' trong sheet ChiTiet PHẢI khớp với slug trong KhachHang", None),
        ("   · Ảnh logo: upload vào assets/omega-media/khach-hang/ trên Git", None),
        ("   · Cache GAS: 5 phút. Dùng menu 📋 OMEGA CMS → Xóa cache để force update ngay", None),
        ("   · URL trang chi tiết tự động tính: khach-hang/index.html?slug=skypec", None),
        ("   · Thêm khách hàng mới: thêm dòng vào KhachHang + ChiTiet, cột Hiển thị = TRUE", None),
    ]
    for i, (txt, bg) in enumerate(lines, 1):
        c = ws.cell(i, 1, txt)
        if bg:
            c.fill = fill(bg)
            c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        else:
            c.font = data_font(size=10)
        c.alignment = left()
        ws.row_dimensions[i].height = 20 if txt else 8

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("Trích xuất dữ liệu từ 19 trang tĩnh...")
    ct_rows = build_chi_tiet_data()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_sheet_kh(wb)
    build_sheet_cs(wb)
    build_sheet_tm(wb)
    build_sheet_ct(wb, ct_rows)
    build_sheet_hd(wb)
    build_sheet_guide(wb)
    build_sheet_icons(wb)

    out = os.path.join(ROOT, "auto-omega", "khach-hang.xlsx")
    wb.save(out)
    print(f"\nDone! Ghi ra: {out}")
    print(f"  · KhachHang  : {len(KH_DATA)} khách hàng")
    print(f"  · CaseStudy  : {len(CS_DATA)} cards")
    print(f"  · Testimonial: {len(TM_DATA)} đánh giá")
    print(f"  · ChiTiet    : {len(ct_rows)} dòng (trích xuất từ trang tĩnh)")
    print(f"  · Guide      : hướng dẫn nhân viên quản trị web")
    print(f"  · Icons      : danh sách icon FA cho dropdown")

if __name__ == "__main__":
    main()
