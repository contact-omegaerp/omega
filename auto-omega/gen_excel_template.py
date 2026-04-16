#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_excel_template.py — Tạo file Excel mẫu cho hệ thống tuyển dụng OMEGA
Chạy: python auto-omega/gen_excel_template.py
"""
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT   = os.path.join(ROOT, 'auto-omega', 'tuyen-dung-sheet-template.xlsx')

# ── Colors ──────────────────────────────────────────────────────────────────
GREEN  = 'FF0d5c38'
LGREEN = 'FFe8f5e9'
YELLOW = 'FFFFF3CD'
LGRAY  = 'FFF5F5F5'
ORANGE = 'FFFF6B35'
RED    = 'FFdc3545'
WHITE  = 'FFFFFFFF'

def hdr_font(bold=True, color='FFFFFFFF', size=11):
    return Font(name='Calibri', bold=bold, color=color, size=size)

def hdr_fill(color):
    return PatternFill('solid', fgColor=color)

def thin_border():
    s = Side(style='thin', color='FFcccccc')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── Sheet 1: Jobs ────────────────────────────────────────────────────────────
def make_jobs_sheet(wb, jobs):
    ws = wb.create_sheet('Jobs')
    ws.freeze_panes = 'A2'

    headers = [
        ('Job_ID', 8, 'ID duy nhất, vd: J001'),
        ('Slug', 38, 'URL slug, vd: tuyen-dung-lap-trinh-vien-backend'),
        ('Title', 40, 'Tiêu đề vị trí tuyển dụng'),
        ('Department', 20, 'Phòng ban'),
        ('Dept_Slug', 20, 'Slug phòng ban, vd: ky-thuat'),
        ('Type', 18, 'Toàn thời gian / Bán thời gian'),
        ('Level', 18, 'Junior, Middle, Senior, Lead, Manager, Director'),
        ('Location', 30, 'Địa điểm làm việc'),
        ('Salary', 25, 'Mức lương, vd: 20 – 35 triệu'),
        ('Openings', 10, 'Số lượng cần tuyển'),
        ('Deadline', 14, 'Hạn nộp hồ sơ dd/mm/yyyy'),
        ('Status', 12, 'ACTIVE | HOT | CLOSED'),
        ('Hot', 8, 'TRUE / FALSE'),
        ('Excerpt', 60, 'Mô tả ngắn 1-2 câu'),
    ]

    # Header row
    for col, (hname, width, _note) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hname)
        cell.font    = hdr_font()
        cell.fill    = hdr_fill(GREEN)
        cell.alignment = center()
        cell.border  = thin_border()
        set_col_width(ws, col, width)

    ws.row_dimensions[1].height = 28

    # Data rows from JOBS constant (inline to avoid import complexity)
    JOBS_DATA = [
        ('J001','tuyen-dung-lap-trinh-vien-backend','Lập trình viên Backend (.NET / Node.js)','Phòng Kỹ thuật','ky-thuat','Toàn thời gian','Senior','TP.HCM','20 – 35 triệu',3,'30/06/2026','ACTIVE',True,'Xây dựng và tối ưu hóa hệ thống backend cho nền tảng ERP OMEGA – xử lý nghiệp vụ phức tạp, hiệu năng cao.'),
        ('J002','tuyen-dung-lap-trinh-vien-frontend','Lập trình viên Frontend (Vue.js / React)','Phòng Kỹ thuật','ky-thuat','Toàn thời gian','Middle – Senior','TP.HCM','18 – 30 triệu',2,'30/06/2026','ACTIVE',True,'Xây dựng giao diện người dùng hiện đại cho hệ thống OMEGA ERP – đẹp, responsive, UX tối ưu.'),
        ('J003','tuyen-dung-lap-trinh-vien-mobile','Lập trình viên Mobile (React Native / Flutter)','Phòng Kỹ thuật','ky-thuat','Toàn thời gian','Middle','TP.HCM','18 – 28 triệu',2,'30/06/2026','ACTIVE',False,'Phát triển ứng dụng mobile cho hệ sinh thái OMEGA (APV, SCR, STK, SOR, MST, HRM App).'),
        ('J004','tuyen-dung-ky-su-qa-testing','Kỹ sư QA / Testing','Phòng Kỹ thuật','ky-thuat','Toàn thời gian','Middle','TP.HCM','15 – 22 triệu',2,'30/06/2026','ACTIVE',False,'Đảm bảo chất lượng sản phẩm OMEGA ERP – xây dựng test plan, viết test case, automation test.'),
        ('J005','tuyen-dung-ky-su-devops','Kỹ sư DevOps / Cloud','Phòng Kỹ thuật','ky-thuat','Toàn thời gian','Senior','TP.HCM','22 – 35 triệu',1,'30/06/2026','ACTIVE',True,'Xây dựng và vận hành hạ tầng cloud, CI/CD pipeline cho hệ thống OMEGA ERP.'),
        ('J006','tuyen-dung-chuyen-gia-ai-automation','Chuyên gia AI & Automation (ERP)','Phòng Kỹ thuật','ky-thuat','Toàn thời gian','Senior','TP.HCM','25 – 45 triệu',1,'30/06/2026','ACTIVE',True,'Nghiên cứu và tích hợp AI (LLM, Gemini, GPT) vào hệ sinh thái OMEGA ERP.'),
        ('J007','tuyen-dung-chuyen-vien-tu-van-trien-khai-erp','Chuyên viên Tư vấn Triển khai ERP','Phòng Triển khai','trien-khai','Toàn thời gian','Middle – Senior','TP.HCM (đi tỉnh theo dự án)','18 – 30 triệu',4,'30/06/2026','ACTIVE',True,'Tư vấn và triển khai OMEGA ERP tại doanh nghiệp khách hàng – phân tích nghiệp vụ, cấu hình, đào tạo.'),
        ('J008','tuyen-dung-chuyen-vien-trien-khai-ke-toan','Chuyên viên Triển khai Kế toán (OMEGA.GL / FA / MC)','Phòng Triển khai','trien-khai','Toàn thời gian','Middle','TP.HCM','16 – 25 triệu',3,'30/06/2026','ACTIVE',False,'Triển khai phân hệ kế toán OMEGA cho doanh nghiệp – cấu hình đến đào tạo kế toán viên.'),
        ('J009','tuyen-dung-chuyen-vien-trien-khai-san-xuat','Chuyên viên Triển khai Sản xuất (OMEGA.MM / PC / QC)','Phòng Triển khai','trien-khai','Toàn thời gian','Middle','TP.HCM (đi tỉnh theo dự án)','16 – 26 triệu',2,'30/06/2026','ACTIVE',False,'Triển khai phân hệ sản xuất OMEGA cho nhà máy – BOM, lệnh sản xuất, kiểm soát chất lượng.'),
        ('J010','tuyen-dung-chuyen-vien-trien-khai-nhan-su','Chuyên viên Triển khai Nhân sự (OMEGA.HR / PR)','Phòng Triển khai','trien-khai','Toàn thời gian','Middle','TP.HCM','15 – 24 triệu',2,'30/06/2026','ACTIVE',False,'Triển khai phân hệ nhân sự và tiền lương OMEGA – cấu hình, đào tạo và hỗ trợ sau go-live.'),
        ('J011','tuyen-dung-business-analyst','Business Analyst (ERP)','Phòng Triển khai','trien-khai','Toàn thời gian','Middle – Senior','TP.HCM','20 – 32 triệu',2,'30/06/2026','ACTIVE',True,'Phân tích nghiệp vụ, xây dựng tài liệu yêu cầu – cầu nối giữa khách hàng và team kỹ thuật.'),
        ('J012','tuyen-dung-chuyen-vien-dao-tao-erp','Chuyên viên Đào tạo ERP','Phòng Triển khai','trien-khai','Toàn thời gian','Junior – Middle','TP.HCM (đi tỉnh theo lịch)','13 – 20 triệu',2,'30/06/2026','ACTIVE',False,'Thiết kế và thực hiện chương trình đào tạo OMEGA ERP cho người dùng cuối.'),
        ('J013','tuyen-dung-truong-nhom-trien-khai','Trưởng nhóm Triển khai ERP','Phòng Triển khai','trien-khai','Toàn thời gian','Leader','TP.HCM','25 – 40 triệu',1,'30/06/2026','ACTIVE',False,'Dẫn dắt nhóm triển khai ERP – quản lý dự án, đảm bảo chất lượng và tiến độ.'),
        ('J014','tuyen-dung-chuyen-vien-kinh-doanh','Chuyên viên Kinh doanh (ERP Software)','Phòng Kinh doanh','kinh-doanh','Toàn thời gian','Junior – Middle','TP.HCM','10 – 15 triệu + hoa hồng không giới hạn',5,'30/06/2026','ACTIVE',True,'Tìm kiếm và phát triển khách hàng doanh nghiệp cho OMEGA ERP – thu nhập không giới hạn.'),
        ('J015','tuyen-dung-chuyen-vien-phat-trien-thi-truong','Chuyên viên Phát triển Thị trường','Phòng Kinh doanh','kinh-doanh','Toàn thời gian','Middle','TP.HCM / Hà Nội','15 – 25 triệu + hoa hồng',2,'30/06/2026','ACTIVE',False,'Mở rộng thị trường OMEGA ERP tại các ngành mục tiêu và khu vực địa lý mới.'),
        ('J016','tuyen-dung-chuyen-vien-digital-marketing','Chuyên viên Digital Marketing','Phòng Kinh doanh','kinh-doanh','Toàn thời gian','Middle','TP.HCM','15 – 22 triệu',1,'30/06/2026','ACTIVE',True,'Triển khai chiến lược marketing số cho OMEGA ERP – tăng nhận diện thương hiệu và tạo lead.'),
        ('J017','tuyen-dung-chuyen-vien-seo-content','Chuyên viên SEO Content','Phòng Kinh doanh','kinh-doanh','Toàn thời gian','Junior – Middle','TP.HCM','12 – 18 triệu',2,'30/06/2026','ACTIVE',False,'Tạo nội dung SEO chất lượng cao cho omega.com.vn – tăng organic traffic từ doanh nghiệp.'),
        ('J018','tuyen-dung-chuyen-vien-pre-sales','Chuyên viên Pre-Sales ERP','Phòng Kinh doanh','kinh-doanh','Toàn thời gian','Senior','TP.HCM','22 – 35 triệu',2,'30/06/2026','ACTIVE',True,'Demo sản phẩm, tư vấn giải pháp và hỗ trợ đội kinh doanh chốt deal ERP.'),
        ('J019','tuyen-dung-chuyen-vien-ho-tro-ky-thuat','Chuyên viên Hỗ trợ Kỹ thuật (Helpdesk ERP)','Phòng Hỗ trợ','ho-tro','Toàn thời gian','Junior – Middle','TP.HCM','12 – 18 triệu',3,'30/06/2026','ACTIVE',False,'Hỗ trợ kỹ thuật cho 1000+ doanh nghiệp dùng OMEGA ERP – xử lý ticket, debug, hướng dẫn.'),
        ('J020','tuyen-dung-chuyen-vien-ho-tro-nghiep-vu','Chuyên viên Hỗ trợ Nghiệp vụ','Phòng Hỗ trợ','ho-tro','Toàn thời gian','Junior – Middle','TP.HCM','12 – 18 triệu',2,'30/06/2026','ACTIVE',False,'Hỗ trợ khách hàng về nghiệp vụ kế toán, nhân sự, sản xuất trong vận hành OMEGA ERP.'),
        ('J021','tuyen-dung-chuyen-vien-cham-soc-khach-hang','Chuyên viên Chăm sóc Khách hàng','Phòng Hỗ trợ','ho-tro','Toàn thời gian','Junior','TP.HCM','10 – 15 triệu',2,'30/06/2026','ACTIVE',False,'Chăm sóc và duy trì mối quan hệ khách hàng dùng OMEGA ERP – tăng retention và upsell.'),
        ('J022','tuyen-dung-quan-ly-du-an-erp','Quản lý Dự án ERP (Project Manager)','Phòng Quản lý','quan-ly','Toàn thời gian','Manager','TP.HCM','28 – 45 triệu',2,'30/06/2026','ACTIVE',True,'Quản lý end-to-end dự án triển khai ERP cho doanh nghiệp – từ kick-off đến go-live.'),
        ('J023','tuyen-dung-truong-nhom-phat-trien','Trưởng nhóm Phát triển (Dev Team Lead)','Phòng Quản lý','quan-ly','Toàn thời gian','Lead','TP.HCM','30 – 50 triệu',1,'30/06/2026','ACTIVE',True,'Dẫn dắt team kỹ thuật phát triển OMEGA ERP – technical leadership, kiến trúc, chất lượng code.'),
        ('J024','tuyen-dung-truong-phong-kinh-doanh','Trưởng phòng Kinh doanh','Phòng Quản lý','quan-ly','Toàn thời gian','Manager','TP.HCM','30 – 50 triệu + hoa hồng nhóm',1,'30/06/2026','ACTIVE',False,'Lãnh đạo phòng kinh doanh OMEGA – chiến lược bán hàng, phát triển đội nhóm, KPI doanh thu.'),
        ('J025','tuyen-dung-truong-phong-ky-thuat','Trưởng phòng Kỹ thuật (CTO / VP Engineering)','Phòng Quản lý','quan-ly','Toàn thời gian','Director','TP.HCM','40 – 65 triệu',1,'30/06/2026','ACTIVE',False,'Định hướng công nghệ và lãnh đạo toàn bộ phòng kỹ thuật OMEGA.'),
        ('J026','tuyen-dung-ke-toan-tong-hop','Kế toán Tổng hợp','Phòng Hành chính','hanh-chinh','Toàn thời gian','Junior – Middle','TP.HCM','12 – 18 triệu',1,'30/06/2026','ACTIVE',False,'Đảm nhận công tác kế toán nội bộ của Công ty TNHH Công nghệ và Giải pháp Omega.'),
        ('J027','tuyen-dung-nhan-vien-hanh-chinh-nhan-su','Nhân viên Hành chính – Nhân sự','Phòng Hành chính','hanh-chinh','Toàn thời gian','Junior','TP.HCM','10 – 14 triệu',1,'30/06/2026','ACTIVE',False,'Hỗ trợ công tác hành chính tổng hợp và nhân sự tại Omega.'),
        ('J028','tuyen-dung-chuyen-vien-tuyen-dung-noi-bo','Chuyên viên Tuyển dụng Nội bộ','Phòng Hành chính','hanh-chinh','Toàn thời gian','Middle','TP.HCM','14 – 20 triệu',1,'30/06/2026','ACTIVE',False,'Chịu trách nhiệm tuyển dụng nhân sự chất lượng cho Omega – IT, triển khai, kinh doanh.'),
        ('J029','tuyen-dung-chuyen-vien-thiet-ke-ui-ux','Chuyên viên Thiết kế UI/UX','Phòng Thiết kế','thiet-ke','Toàn thời gian','Middle','TP.HCM','18 – 28 triệu',1,'30/06/2026','ACTIVE',True,'Thiết kế giao diện người dùng cho OMEGA ERP và ứng dụng mobile – UX thông minh, UI sạch.'),
        ('J030','tuyen-dung-chuyen-vien-thiet-ke-do-hoa','Chuyên viên Thiết kế Đồ họa','Phòng Thiết kế','thiet-ke','Toàn thời gian','Junior – Middle','TP.HCM','12 – 20 triệu',1,'30/06/2026','ACTIVE',False,'Thiết kế ấn phẩm truyền thông, marketing và thương hiệu cho OMEGA – digital đến print.'),
    ]

    status_fill = {'ACTIVE': hdr_fill('FF28a745'), 'HOT': hdr_fill('FFdc3545'), 'CLOSED': hdr_fill('FF6c757d')}

    for row_i, row in enumerate(JOBS_DATA, 2):
        fill = hdr_fill(LGREEN) if row_i % 2 == 0 else hdr_fill(WHITE)
        for col_i, val in enumerate(row, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = thin_border()
            cell.alignment = left() if col_i > 3 else center()
            cell.font = Font(name='Calibri', size=10)
            # Status column (col 12) colored
            if col_i == 12:
                s = status_fill.get(val, hdr_fill(WHITE))
                cell.fill = s
                cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
                cell.alignment = center()
            elif col_i == 1:
                cell.font = Font(name='Calibri', size=10, bold=True, color='FF0d5c38')
            else:
                cell.fill = fill
        ws.row_dimensions[row_i].height = 22

    # Add note row
    note = ws.cell(row=len(JOBS_DATA)+3, column=1, value='⚠ Lưu ý: Cột Status chỉ nhận giá trị ACTIVE | HOT | CLOSED. Sau khi upload lên Google Sheets, điền SHEET_ID vào tuyen-dung.gs.')
    note.font = Font(name='Calibri', italic=True, color='FF856404', size=10)
    note.alignment = left()
    ws.merge_cells(f'A{len(JOBS_DATA)+3}:N{len(JOBS_DATA)+3}')

    ws.sheet_view.showGridLines = True
    print(f'  Jobs sheet: {len(JOBS_DATA)} rows')
    return ws


# ── Sheet 2: Applications ────────────────────────────────────────────────────
def make_applications_sheet(wb):
    ws = wb.create_sheet('Applications')
    ws.freeze_panes = 'A2'

    headers = [
        ('App_ID', 14, 'Auto-generated, vd: APP-20260416-001'),
        ('Timestamp', 22, 'Thời gian nộp hồ sơ'),
        ('Job_ID', 10, 'ID vị trí ứng tuyển'),
        ('Job_Title', 40, 'Tên vị trí ứng tuyển'),
        ('Full_Name', 25, 'Họ và tên ứng viên'),
        ('Email', 30, 'Email ứng viên'),
        ('Phone', 16, 'Số điện thoại'),
        ('CV_Link', 50, 'Link Google Drive hoặc LinkedIn'),
        ('CV_Text', 60, 'Nội dung CV paste trực tiếp'),
        ('AI_Score', 10, 'Điểm AI đánh giá (0-100)'),
        ('AI_Summary', 80, 'Nhận xét tổng hợp của AI'),
        ('AI_Strengths', 60, 'Điểm mạnh theo AI'),
        ('AI_Gaps', 60, 'Điểm cần cải thiện theo AI'),
        ('AI_Recommendation', 20, 'Khuyến nghị: HIRE | CONSIDER | REJECT'),
        ('HR_Status', 18, 'Trạng thái HR: New | Reviewing | Interview | Offer | Rejected'),
        ('HR_Note', 60, 'Ghi chú HR'),
    ]

    for col, (hname, width, note) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hname)
        cell.font = hdr_font()
        cell.fill = hdr_fill('FF1565C0')
        cell.alignment = center()
        cell.border = thin_border()
        set_col_width(ws, col, width)

    ws.row_dimensions[1].height = 28

    # Sample data row (template)
    sample = ['APP-20260416-001','16/04/2026 09:30','J001','Lập trình viên Backend (.NET / Node.js)',
              'Nguyễn Văn A','nguyenvana@email.com','0909123456','https://drive.google.com/...',
              '(hoặc nội dung CV dán tại đây)','82',
              'Ứng viên có kinh nghiệm tốt, phù hợp 82% với yêu cầu vị trí.',
              'Kinh nghiệm .NET 4 năm, có dự án ERP thực tế',
              'Chưa có kinh nghiệm với SQL Server nâng cao',
              'CONSIDER','New','Hẹn phỏng vấn tuần tới']
    for col, val in enumerate(sample, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = hdr_fill(YELLOW)
        cell.font = Font(name='Calibri', size=10, italic=True, color='FF856404')
        cell.alignment = left()
        cell.border = thin_border()
    ws.row_dimensions[2].height = 22

    note = ws.cell(row=4, column=1, value='⚠ Sheet này được điền TỰ ĐỘNG bởi Apps Script (tuyen-dung.gs) khi ứng viên nộp hồ sơ. Không cần sửa thủ công.')
    note.font = Font(name='Calibri', italic=True, color='FF856404', size=10)
    note.alignment = left()
    ws.merge_cells('A4:P4')

    print('  Applications sheet: created')


# ── Sheet 3: Config ──────────────────────────────────────────────────────────
def make_config_sheet(wb):
    ws = wb.create_sheet('Config')
    ws.freeze_panes = 'A2'

    # Header
    for col, (h, w) in enumerate([('Key', 28), ('Value', 60), ('Description', 60)], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font()
        cell.fill = hdr_fill('FF37474F')
        cell.alignment = center()
        cell.border = thin_border()
        set_col_width(ws, col, w)
    ws.row_dimensions[1].height = 28

    configs = [
        # Section: LLM
        ('── LLM Configuration ──', '', ''),
        ('LLM_Provider', 'Gemini', 'Nhà cung cấp AI: Gemini | DeepSeek | Grok'),
        ('LLM_Model', 'gemini-2.0-flash', 'Tên model AI (PRIVATE – không được trả về frontend)'),
        ('LLM_API_Key', '', 'API Key của nhà cung cấp AI (PRIVATE – không trả về frontend)'),
        ('LLM_Max_Token', '2048', 'Số token tối đa cho phản hồi AI'),
        # Section: Email
        ('── Email Configuration ──', '', ''),
        ('HR_Emails', 'hr@omega.com.vn;tuyen-dung@omega.com.vn', 'Email HR nhận thông báo – phân cách bằng dấu ; (PRIVATE)'),
        ('Email_From_Name', 'OMEGA HR Team', 'Tên hiển thị trong email gửi ra'),
        ('Email_Reply_To', 'hr@omega.com.vn', 'Email reply-to cho ứng viên'),
        # Section: System
        ('── System Settings ──', '', ''),
        ('Company_Name', 'Công ty TNHH Công nghệ và Giải pháp Omega', 'Tên công ty đầy đủ'),
        ('Company_Name_EN', 'Omega Solution & Technology Co., Ltd.', 'Tên tiếng Anh'),
        ('Company_Website', 'https://omega.com.vn', 'Website chính'),
        ('Min_Score_Pass', '60', 'Điểm AI tối thiểu để tự động chuyển trạng thái CONSIDER'),
        ('Auto_Reply_Delay_Hours', '24', 'Thời gian tối đa gửi email phản hồi cho ứng viên (giờ)'),
        ('Max_CV_Text_Length', '10000', 'Giới hạn ký tự cho CV Text paste'),
        # Section: Security
        ('── Security (PRIVATE) ──', '', ''),
        ('Master_Password', '', 'Mật khẩu master để reset/override (PRIVATE – không trả về frontend)'),
        ('Allowed_Origins', 'https://omega.com.vn', 'Các domain được phép POST (CORS)'),
    ]

    section_fill   = hdr_fill('FF37474F')
    private_fill   = hdr_fill('FFFFEBEE')
    section_keys   = ['── LLM Configuration ──', '── Email Configuration ──', '── System Settings ──', '── Security (PRIVATE) ──']
    private_keys   = ['LLM_Model', 'LLM_API_Key', 'HR_Emails', 'Master_Password']

    for row_i, (key, val, desc) in enumerate(configs, 2):
        is_section = key in section_keys
        is_private = key in private_keys

        for col_i, v in enumerate([key, val, desc], 1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.border = thin_border()
            cell.alignment = left()
            if is_section:
                cell.font = Font(name='Calibri', bold=True, color='FFFFFFFF', size=10)
                cell.fill = hdr_fill('FF546E7A')
                cell.alignment = center()
            elif is_private:
                cell.font = Font(name='Calibri', size=10, color='FFCC0000', bold=(col_i == 1))
                cell.fill = private_fill
            else:
                cell.font = Font(name='Calibri', size=10, color='FF1B5E20' if col_i == 1 else 'FF212121')
                cell.fill = hdr_fill(LGREEN) if row_i % 2 == 0 else hdr_fill(WHITE)

        ws.row_dimensions[row_i].height = 20

    # Legend
    leg_row = len(configs) + 3
    legend_data = [
        (leg_row,   'FF800000', '🔒 Các dòng màu đỏ nhạt là PRIVATE KEYS – Apps Script sẽ KHÔNG trả về các giá trị này cho frontend HTML.'),
        (leg_row+1, 'FF0d5c38', '✅ Sau khi upload file này lên Google Sheets, copy Sheet ID từ URL và điền vào hằng số SHEET_ID trong tuyen-dung.gs.'),
        (leg_row+2, 'FF1565C0', 'ℹ️ URL Sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit'),
    ]
    for lrow, color, text in legend_data:
        cell = ws.cell(row=lrow, column=1, value=text)
        cell.font = Font(name='Calibri', italic=True, color=color, size=10)
        cell.alignment = left()
        ws.merge_cells(f'A{lrow}:C{lrow}')

    print('  Config sheet: created')


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    make_jobs_sheet(wb, None)
    make_applications_sheet(wb)
    make_config_sheet(wb)

    # Tab colors
    wb['Jobs'].sheet_properties.tabColor = '0d5c38'
    wb['Applications'].sheet_properties.tabColor = '1565C0'
    wb['Config'].sheet_properties.tabColor = '37474F'

    wb.save(OUT)
    print(f'\n  Saved: {OUT}')
    print('  Sheets: Jobs | Applications | Config')
    print('\nHướng dẫn:')
    print('  1. Upload file này lên Google Drive > "Save as Google Sheets"')
    print('  2. Copy Sheet ID từ URL: https://docs.google.com/spreadsheets/d/{ID}/edit')
    print('  3. Điền ID vào const SHEET_ID trong auto-omega/tuyen-dung.gs')
    print('  4. Điền LLM_API_Key và HR_Emails vào sheet Config')
    print('  5. Deploy tuyen-dung.gs > copy Deployment URL vào biến GAS_URL trong HTML')


if __name__ == '__main__':
    main()
